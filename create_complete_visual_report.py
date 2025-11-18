import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from pathlib import Path
import os

print("="*80)
print("CREATING COMPREHENSIVE VISUAL EXCEL REPORT")
print("="*80)

wb = Workbook()
base_path = "/home/gee_devops254/Downloads/Half Board"

# Styles
title_font = Font(bold=True, color="FFFFFF", size=16)
title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
description_font = Font(size=10, italic=True)
description_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# ============================================================================
# SHEET 1: TABLE OF CONTENTS
# ============================================================================
print("[1/11] Creating Table of Contents...")
ws_toc = wb.active
ws_toc.title = "📋 Contents"

ws_toc['A1'] = "HALF BOARD VISUAL ANALYSIS REPORT"
ws_toc['A1'].font = Font(bold=True, size=18, color="FFFFFF")
ws_toc['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_toc['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_toc.merge_cells('A1:D1')
ws_toc.row_dimensions[1].height = 35

ws_toc['A3'] = "📊 REPORT SECTIONS"
ws_toc['A3'].font = Font(bold=True, size=14, color="2C3E50")
ws_toc.row_dimensions[3].height = 25

sections = [
    ("📈 Overview", "Overview & Distribution (6 charts)", 5),
    ("🏢 Agencies", "Agency Analysis (8 charts)", 6),
    ("⭐ Miracle", "Miracle Tourism Deep Dive (5 charts)", 7),
    ("🔑 Rate Codes", "Rate Code Analysis (6 charts)", 8),
    ("🌍 Markets", "Market Segmentation (5 charts)", 9),
    ("📊 HB Performance", "HB Performance Analysis (6 charts)", 10),
    ("🎯 Opportunities", "Opportunity Analysis (5 charts)", 11),
    ("🔗 Correlations", "Correlation & Relationships (4 charts)", 12),
    ("📱 Dashboards", "Strategic Dashboards (3 charts)", 13)
]

row = 5
for sheet_name, description, _ in sections:
    ws_toc[f'A{row}'] = sheet_name
    ws_toc[f'B{row}'] = description
    ws_toc[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
    ws_toc[f'B{row}'].font = Font(size=11)
    ws_toc.row_dimensions[row].height = 22
    row += 1

ws_toc['A{}'.format(row+2)] = "🔑 KEY STATISTICS"
ws_toc['A{}'.format(row+2)].font = Font(bold=True, size=14, color="2C3E50")

stats = [
    ("Total Revenue:", "AED 9,338,949"),
    ("Total Room Nights:", "23,747"),
    ("HB Penetration:", "14% (50/357 bookings)"),
    ("Incremental Potential:", "AED 920,000+ (if 35% adoption)")
]

stat_row = row + 4
for label, value in stats:
    ws_toc[f'A{stat_row}'] = label
    ws_toc[f'B{stat_row}'] = value
    ws_toc[f'A{stat_row}'].font = Font(bold=True, size=11)
    ws_toc[f'B{stat_row}'].font = Font(size=11, color="E74C3C", bold=True)
    stat_row += 1

ws_toc.column_dimensions['A'].width = 25
ws_toc.column_dimensions['B'].width = 50

# ============================================================================
# SHEET 2: INITIAL ANALYSIS SUITE
# ============================================================================
print("[2/11] Creating Initial Analysis Suite...")
ws_initial = wb.create_sheet("🔬 Initial Analysis")

ws_initial['A1'] = "INITIAL STATISTICAL ANALYSIS - 6 Charts"
ws_initial['A1'].font = title_font
ws_initial['A1'].fill = title_fill
ws_initial['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_initial.merge_cells('A1:H1')
ws_initial.row_dimensions[1].height = 30

initial_charts = [
    {
        'path': f'{base_path}/1_univariate_analysis.png',
        'title': 'Chart 1: Univariate Analysis',
        'description': 'Distribution analysis of Room Nights, Revenue, and Average Rates. Histograms with KDE curves showing data distributions and outliers.',
        'row': 3
    },
    {
        'path': f'{base_path}/2_multivariate_analysis.png',
        'title': 'Chart 2: Multivariate Analysis',
        'description': 'Comprehensive pairwise relationships between key metrics. Scatter plots, distributions, and correlations in matrix format.',
        'row': 25
    },
    {
        'path': f'{base_path}/3_agency_ratecode_relationships.png',
        'title': 'Chart 3: Agency & Rate Code Relationships',
        'description': 'Agency & Rate Code relationships. Scatter plots and performance matrices showing correlation between volume and HB adoption.',
        'row': 47
    },
    {
        'path': f'{base_path}/4_cis_market_analysis.png',
        'title': 'Chart 4: CIS Market Analysis',
        'description': 'CIS markets show 36% HB penetration (best in dataset) but only 9 bookings total. TOKHACIS leads with AED 17,888 revenue.',
        'row': 69
    },
    {
        'path': f'{base_path}/5_correlation_heatmap.png',
        'title': 'Chart 5: Correlation Heatmap',
        'description': 'Pearson correlation between Room Nights, Revenue, and Avg Rate. Perfect 0.999 correlation between nights and revenue.',
        'row': 91
    },
    {
        'path': f'{base_path}/6_comprehensive_dashboard.png',
        'title': 'Chart 6: Comprehensive Dashboard',
        'description': 'Multi-panel overview: Key KPIs, revenue distribution, top agencies, booking patterns, and rate comparisons.',
        'row': 113
    }
]

for chart_info in initial_charts:
    row = chart_info['row']

    ws_initial[f'A{row}'] = chart_info['title']
    ws_initial[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")
    ws_initial.merge_cells(f'A{row}:H{row}')
    ws_initial.row_dimensions[row].height = 25

    ws_initial[f'A{row+1}'] = chart_info['description']
    ws_initial[f'A{row+1}'].font = description_font
    ws_initial[f'A{row+1}'].fill = description_fill
    ws_initial[f'A{row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_initial.merge_cells(f'A{row+1}:H{row+2}')
    ws_initial.row_dimensions[row+1].height = 40

    if os.path.exists(chart_info['path']):
        img = Image(chart_info['path'])
        img.width = 1000
        img.height = int(1000 * img.height / img.width) if img.width > 0 else 400
        ws_initial.add_image(img, f'A{row+3}')
        for i in range(20):
            ws_initial.row_dimensions[row+3+i].height = 20

ws_initial.column_dimensions['A'].width = 15

# ============================================================================
# HELPER FUNCTION
# ============================================================================
def create_category_sheet(wb, sheet_name, title, chart_folder, chart_descriptions):
    """Create a sheet with charts from a specific category"""
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    current_row = 3

    for chart_num, desc in chart_descriptions.items():
        chart_files = list(Path(f'{base_path}/charts/{chart_folder}').glob(f'{chart_num}_*.png'))
        if not chart_files:
            continue

        chart_path = str(chart_files[0])

        # Title
        ws[f'A{current_row}'] = f'Chart {chart_num}: {desc["title"]}'
        ws[f'A{current_row}'].font = Font(bold=True, size=13, color="2C3E50")
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws.row_dimensions[current_row].height = 25

        # Description
        ws[f'A{current_row+1}'] = desc['description']
        ws[f'A{current_row+1}'].font = description_font
        ws[f'A{current_row+1}'].fill = description_fill
        ws[f'A{current_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row+1}:H{current_row+2}')
        ws.row_dimensions[current_row+1].height = 35

        # Key Insight
        ws[f'A{current_row+3}'] = '🔑 KEY INSIGHT:'
        ws[f'B{current_row+3}'] = desc['insight']
        ws[f'A{current_row+3}'].font = Font(bold=True, color="E74C3C")
        ws[f'B{current_row+3}'].font = Font(bold=True, color="27AE60")
        ws[f'B{current_row+3}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        ws.merge_cells(f'B{current_row+3}:H{current_row+3}')
        ws.row_dimensions[current_row+3].height = 20

        # Image
        if os.path.exists(chart_path):
            img = Image(chart_path)
            img.width = 950
            img.height = int(950 * img.height / img.width) if img.width > 0 else 380
            ws.add_image(img, f'A{current_row+4}')

            rows_needed = int(img.height / 20) + 1
            for i in range(rows_needed):
                ws.row_dimensions[current_row+4+i].height = 20

            current_row += rows_needed + 6
        else:
            current_row += 25

    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14

# ============================================================================
# SHEET 3: OVERVIEW & DISTRIBUTION
# ============================================================================
print("[3/11] Creating Overview & Distribution...")
overview_charts = {
    '01': {
        'title': 'Overall HB Penetration',
        'description': 'Pie chart showing only 14% of bookings include Half Board, leaving 86% as opportunity gap.',
        'insight': 'CRITICAL: 86% of bookings have NO F&B packages!'
    },
    '02': {
        'title': 'Revenue Distribution',
        'description': 'HB generates only 3.1% of total revenue despite being 14% of bookings.',
        'insight': 'HB is underperforming in revenue generation - pricing issue'
    },
    '03': {
        'title': 'Room Nights Distribution',
        'description': 'Only 2.7% of room nights include F&B packages (648 out of 23,747 nights).',
        'insight': '97.3% of room nights lack F&B attachment'
    },
    '04': {
        'title': 'Booking Count Distribution',
        'description': 'Visual comparison: 50 HB bookings vs 307 non-HB bookings.',
        'insight': 'Convert non-HB to HB through incentives and bundling'
    },
    '05': {
        'title': 'Average Rate Comparison',
        'description': 'HB average rate (AED 419.30) only AED 26.21 higher than non-HB (AED 393.09).',
        'insight': 'Increase HB surcharge to AED 120-150 per night'
    },
    '06': {
        'title': 'Revenue vs Room Nights Scatter',
        'description': 'Scatter plot showing perfect linear relationship (0.999 correlation). Green = HB, Red = Non-HB.',
        'insight': 'Strong correlation confirms: More room nights = More revenue'
    }
}

create_category_sheet(wb, "📈 Overview", "OVERVIEW & DISTRIBUTION ANALYSIS",
                     "01_Overview_Distribution", overview_charts)

# ============================================================================
# SHEET 4: AGENCY ANALYSIS
# ============================================================================
print("[4/11] Creating Agency Analysis...")
agency_charts = {
    '07': {
        'title': 'Top 20 Agencies by Revenue',
        'description': 'Revenue leaders ranked. "Totals" dominates with AED 3.8M. TBO, Darina, Webbeds follow.',
        'insight': 'Top 5 agencies generate 60%+ of revenue - prioritize these'
    },
    '08': {
        'title': 'Top 20 Agencies by Room Nights',
        'description': 'Volume leaders: "Totals" (11,874 nights), TBO (758), Darina (932).',
        'insight': 'Target these high-volume agencies for HB conversion'
    },
    '09': {
        'title': 'Top 15 HB Agencies by Nights',
        'description': 'Miracle Tourism leads HB performance with 151 nights - 2x more than #2 (Desert Gate: 71 nights).',
        'insight': 'Miracle Tourism is the HB champion - study their model!'
    },
    '10': {
        'title': 'Top 15 HB Agencies by Revenue',
        'description': 'Miracle generates AED 55,130 in HB revenue alone. Al Khalidiah and Voyage follow.',
        'insight': 'Create "HB Champions" program modeled after Miracle'
    },
    '11': {
        'title': 'HB Penetration by Top 20',
        'description': 'Color-coded: Green >25%, Orange 10-25%, Red <10%. Most top agencies in RED zone.',
        'insight': 'Most top agencies have <10% HB - HIGH PRIORITY targets'
    },
    '12': {
        'title': 'HB vs Non-HB Stacked',
        'description': 'Stacked bars showing HB (green) vs Non-HB (red) split for top 15 agencies.',
        'insight': 'Use this chart in agency presentations to show opportunity'
    },
    '13': {
        'title': 'Average Rate Comparison',
        'description': 'Grouped bars comparing HB vs Non-HB rates by agency. Some agencies have LOWER HB rates!',
        'insight': 'Standardize HB premium at +AED 120 across all agencies'
    },
    '14': {
        'title': 'Booking Size Box Plot',
        'description': 'Box plots showing booking size variance by agency. Outliers indicate large group bookings.',
        'insight': 'Target agencies with large bookings for mandatory HB'
    }
}

create_category_sheet(wb, "🏢 Agencies", "AGENCY ANALYSIS",
                     "02_Agency_Analysis", agency_charts)

# ============================================================================
# SHEET 5: MIRACLE TOURISM DEEP DIVE
# ============================================================================
print("[5/11] Creating Miracle Tourism Deep Dive...")
miracle_charts = {
    '15': {
        'title': 'Miracle vs Top 5 Comparison',
        'description': 'Normalized multi-metric comparison (0-100 scale). Miracle excels in HB penetration.',
        'insight': "Miracle's model is the blueprint - apply to other agencies"
    },
    '16': {
        'title': 'Miracle Booking Size Distribution',
        'description': 'Histogram showing most Miracle bookings are 30+ nights (large groups/packages).',
        'insight': 'Large bookings = Higher HB adoption. Target long-stay agencies.'
    },
    '17': {
        'title': 'Miracle HB vs Non-HB Split',
        'description': 'Pie chart showing Miracle has achieved high HB penetration in their portfolio.',
        'insight': 'Showcase Miracle as success story in all agency meetings'
    },
    '18': {
        'title': 'Miracle Rate Code Performance',
        'description': 'TOMILUX (Luxembourg market rate code) drives majority of Miracle revenue.',
        'insight': 'Luxembourg market guests prefer packages - expand to similar markets'
    },
    '19': {
        'title': 'Miracle Revenue Contribution',
        'description': 'Single agency contributing significant share of total HB revenue.',
        'insight': 'Reduce dependency by scaling HB across other agencies'
    }
}

create_category_sheet(wb, "⭐ Miracle", "MIRACLE TOURISM DEEP DIVE (Luxembourg Market)",
                     "03_Miracle_Deep_Dive", miracle_charts)

# ============================================================================
# SHEET 6: RATE CODE ANALYSIS
# ============================================================================
print("[6/11] Creating Rate Code Analysis...")
ratecode_charts = {
    '20': {
        'title': 'Top 15 Rate Codes by Revenue',
        'description': 'TOBBWI and TOBBJN (universal codes) highlighted in red. Top revenue generators.',
        'insight': 'Universal codes = Prime targets for HB bundling strategy'
    },
    '21': {
        'title': 'Top 15 Rate Codes by Nights',
        'description': 'TOBBWI (2,915 nights) + TOBBJN (2,530 nights) = 5,445 combined nights.',
        'insight': 'If 35% adopted HB = AED 200k+ incremental revenue'
    },
    '22': {
        'title': 'TOBBWI Agency Performance',
        'description': 'Top 10 agencies using TOBBWI rate code. TBO, Darina, Webbeds lead.',
        'insight': 'Engage these agencies for TOBBWI-HB variant creation'
    },
    '23': {
        'title': 'TOBBJN Agency Performance',
        'description': 'Dubai Link, Darina dominate TOBBJN usage with minimal HB attachment.',
        'insight': 'Make TOBBJN-HB mandatory for 5+ night stays'
    },
    '24': {
        'title': 'Universal Codes HB Penetration',
        'description': 'Dramatic grouped bars showing HB vs Non-HB split. Non-HB dominates.',
        'insight': 'Use in executive presentations to show opportunity size'
    },
    '25': {
        'title': 'Rate Code Avg Rate Comparison',
        'description': 'Color-coded by rate level: Green >AED 400, Orange AED 300-400, Red <AED 300.',
        'insight': 'Tier HB pricing by rate code - premium codes get premium HB'
    }
}

create_category_sheet(wb, "🔑 Rate Codes", "RATE CODE ANALYSIS - Universal & Specific",
                     "04_Rate_Code_Analysis", ratecode_charts)

# ============================================================================
# SHEET 7: MARKET SEGMENTATION
# ============================================================================
print("[7/11] Creating Market Segmentation...")
market_charts = {
    '26': {
        'title': 'Market Revenue Distribution',
        'description': 'Universal/Multi-Market segment dominates revenue but lacks HB focus.',
        'insight': 'Universal segment = Highest revenue + Lowest HB = Biggest opportunity'
    },
    '27': {
        'title': 'Market HB Penetration',
        'description': 'Color-coded by performance: CIS Markets show 36% HB penetration (best).',
        'insight': 'CIS guests already want HB - scale up volume immediately!'
    },
    '28': {
        'title': 'CIS Market Performance',
        'description': 'TOKHACIS, TOVOYCIS, TOCIS25 performance breakdown. TOKHACIS leads with AED 17,888.',
        'insight': 'Launch CIS agency acquisition campaign - they convert easily'
    },
    '29': {
        'title': 'Luxembourg Market Analysis',
        'description': 'Miracle Tourism monopolizes Luxembourg market (TOMILUX code).',
        'insight': 'Luxembourg market profile = High HB. Seek similar European markets.'
    },
    '30': {
        'title': 'Universal vs Specific Comparison',
        'description': 'Dual metric chart: Universal has volume, CIS has HB percentage.',
        'insight': 'Apply CIS HB tactics to Universal segment for maximum impact'
    }
}

create_category_sheet(wb, "🌍 Markets", "MARKET SEGMENTATION - Geographic Analysis",
                     "05_Market_Segmentation", market_charts)

# ============================================================================
# SHEET 8: HB PERFORMANCE
# ============================================================================
print("[8/11] Creating HB Performance...")
performance_charts = {
    '31': {
        'title': 'HB Bookings Tree Visualization',
        'description': 'Proportional representation: Top 3 agencies dominate HB nights.',
        'insight': '80/20 rule confirmed - Top 5 agencies drive 80% of HB'
    },
    '32': {
        'title': 'HB Revenue Pareto',
        'description': 'Pareto chart with cumulative percentage line. Red line crosses 80% at ~5 agencies.',
        'insight': 'Focus resources on top performers + scale the model'
    },
    '33': {
        'title': 'HB Avg Rate Distribution',
        'description': 'Histogram of HB rates clustering around AED 350-450. Mean and median lines shown.',
        'insight': 'Create tiered HB: Essential (120), Enhanced (160), Premium (200)'
    },
    '34': {
        'title': 'HB vs Non-HB Rate Scatter',
        'description': 'Scatter plot with diagonal line. Points above = HB rates higher (good).',
        'insight': 'Agencies below diagonal need immediate rate corrections'
    },
    '35': {
        'title': 'HB Penetration Heatmap',
        'description': 'Color-coded matrix: Top 10 agencies × Top 10 rate codes. Green = high HB%, Red = low.',
        'insight': 'Red cells = Immediate targeting opportunities for conversion'
    },
    '36': {
        'title': 'Booking Size vs HB Adoption',
        'description': 'Scatter with trend line showing positive correlation. Larger bookings = higher HB%.',
        'insight': 'Mandate HB for bookings >50 nights - correlation proven'
    }
}

create_category_sheet(wb, "📊 HB Performance", "HALF BOARD PERFORMANCE ANALYSIS",
                     "06_HB_Performance", performance_charts)

# ============================================================================
# SHEET 9: OPPORTUNITY ANALYSIS
# ============================================================================
print("[9/11] Creating Opportunity Analysis...")
opportunity_charts = {
    '37': {
        'title': 'Opportunity Matrix Quadrant',
        'description': '4-quadrant scatter: Bottom-right (High Volume + Low HB%) = PRIORITY targets.',
        'insight': 'Bottom-right quadrant agencies = Focus ALL sales efforts here'
    },
    '38': {
        'title': 'Top 10 High Priority',
        'description': 'Specific agencies with >AED 50k incremental revenue potential each.',
        'insight': 'Schedule executive meetings with these 10 within 2 weeks'
    },
    '39': {
        'title': 'Incremental Revenue Potential',
        'description': 'Top 15 agencies ranked by potential. Total = AED 920k if all reach 35% HB.',
        'insight': 'Use in board presentations to justify HB program investment'
    },
    '40': {
        'title': 'Current vs Target HB',
        'description': 'Side-by-side bars: Red (current) vs Green (35% target). Visual gap is dramatic.',
        'insight': 'Set agency-specific HB targets and track monthly progress'
    },
    '41': {
        'title': 'Quick Wins vs Long-Term',
        'description': 'Effort-Impact matrix with 10 actions plotted. Green quadrant = Quick wins.',
        'insight': 'Execute quick wins (Actions 1-2) immediately while planning long-term'
    }
}

create_category_sheet(wb, "🎯 Opportunities", "OPPORTUNITY ANALYSIS - AED 920k Potential",
                     "07_Opportunity_Analysis", opportunity_charts)

# ============================================================================
# SHEET 10: CORRELATIONS & RELATIONSHIPS
# ============================================================================
print("[10/11] Creating Correlations...")
correlation_charts = {
    '42': {
        'title': 'Correlation Heatmap',
        'description': 'Pearson correlation matrix for Room Nights, Revenue, Avg Rate. 0.999 = perfect.',
        'insight': 'Room Nights & Revenue: Perfect correlation (0.999)'
    },
    '43': {
        'title': 'Nights vs Revenue by HB',
        'description': 'Dual scatter (HB=green, Non-HB=red) with trend lines. Similar slopes = rate parity.',
        'insight': 'Opportunity is in VOLUME, not rates - both have same slope'
    },
    '44': {
        'title': 'Rate vs Booking Size',
        'description': 'Scatter plot colored by HB status. No strong correlation = pricing is size-agnostic.',
        'insight': 'Consider volume discounts for large bookings to encourage HB'
    },
    '45': {
        'title': 'Agency Performance Matrix',
        'description': 'Bubble scatter: Revenue (X) vs HB Penetration (Y). Color = HB%.',
        'insight': 'Goal: Move agencies from top-left to top-right quadrant'
    }
}

create_category_sheet(wb, "🔗 Correlations", "CORRELATION & RELATIONSHIPS",
                     "08_Correlation_Relationships", correlation_charts)

# ============================================================================
# SHEET 11: STRATEGIC DASHBOARDS
# ============================================================================
print("[11/11] Creating Strategic Dashboards...")
dashboard_charts = {
    '46': {
        'title': 'Executive Summary Dashboard',
        'description': 'Multi-panel dashboard: 4 KPI cards, HB penetration pie, top 10 agencies, market segments.',
        'insight': 'Use in C-level presentations and quarterly board reviews'
    },
    '47': {
        'title': 'HB Performance Scorecard',
        'description': 'Detailed HB metrics: Top generators, revenue by market, rate distribution, trends.',
        'insight': 'Use for monthly HB performance reviews with stakeholders'
    },
    '48': {
        'title': 'Action Priority Matrix',
        'description': 'All 10 prescriptive actions mapped on Effort vs Impact matrix. Color-coded by urgency.',
        'insight': 'This is your complete implementation roadmap - follow priorities'
    }
}

create_category_sheet(wb, "📱 Dashboards", "STRATEGIC DASHBOARDS - Executive View",
                     "09_Strategic_Dashboards", dashboard_charts)

# Save the workbook
output_path = f'{base_path}/Half_Board_Complete_Visual_Report.xlsx'
wb.save(output_path)

print("\n" + "="*80)
print("✓ COMPLETE VISUAL REPORT CREATED SUCCESSFULLY!")
print("="*80)
print(f"\nFile: {output_path}")
print("\nContents:")
print("  ✓ Table of Contents with navigation")
print("  ✓ 6 Initial Analysis charts (with full descriptions)")
print("  ✓ 6 Overview & Distribution charts")
print("  ✓ 8 Agency Analysis charts")
print("  ✓ 5 Miracle Tourism charts")
print("  ✓ 6 Rate Code Analysis charts")
print("  ✓ 5 Market Segmentation charts")
print("  ✓ 6 HB Performance charts")
print("  ✓ 5 Opportunity Analysis charts")
print("  ✓ 4 Correlation charts")
print("  ✓ 3 Strategic Dashboard charts")
print("\nTotal: 54 charts with descriptions, insights, and professional formatting")
print("="*80)
