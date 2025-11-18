import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
import warnings
warnings.filterwarnings('ignore')

print("="*80)
print("COMPREHENSIVE HALF BOARD DEEP-DIVE ANALYSIS (REVISED)")
print("="*80)

# Read the data
df = pd.read_excel('/home/gee_devops254/Downloads/Half Board/Half Board.xlsx')

# Data preparation
df['Avg_Rate_Per_Night'] = df['Room Revenue'] / df['Room Nights']
df['Has_HB'] = df['Product (Descriptions)'].str.contains('Halfboard|Half Board', case=False, na=False)

print(f"\nTotal records: {len(df)}")
print(f"Records with HB: {df['Has_HB'].sum()}")
print(f"HB Penetration: {df['Has_HB'].sum()/len(df)*100:.1f}%")

# ============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================================
print("\n[1/10] Building Executive Summary...")

exec_summary = {
    'Metric': [],
    'Value': [],
    'Notes': []
}

exec_summary['Metric'].extend([
    'Total Bookings',
    'Total Room Nights',
    'Total Revenue (AED)',
    'Average Rate (AED)',
    '',
    'Half Board Bookings',
    'HB Room Nights',
    'HB Revenue (AED)',
    'HB Average Rate (AED)',
    '',
    'HB Penetration Rate (%)',
    'HB Revenue Contribution (%)',
    'HB Room Night Share (%)',
    '',
    'Total Unique Agencies',
    'Agencies with HB',
    'Agencies without HB',
    '',
    'Top Rate Codes',
    'TOBBWI Room Nights',
    'TOBBJN Room Nights',
])

df_hb = df[df['Has_HB']]
df_non_hb = df[~df['Has_HB']]

exec_summary['Value'].extend([
    len(df),
    df['Room Nights'].sum(),
    f"AED {df['Room Revenue'].sum():,.2f}",
    f"AED {df['Avg_Rate_Per_Night'].mean():.2f}",
    '',
    len(df_hb),
    df_hb['Room Nights'].sum(),
    f"AED {df_hb['Room Revenue'].sum():,.2f}",
    f"AED {df_hb['Avg_Rate_Per_Night'].mean():.2f}",
    '',
    f"{len(df_hb)/len(df)*100:.1f}%",
    f"{df_hb['Room Revenue'].sum()/df['Room Revenue'].sum()*100:.1f}%",
    f"{df_hb['Room Nights'].sum()/df['Room Nights'].sum()*100:.1f}%",
    '',
    df['Search Name'].nunique(),
    df_hb['Search Name'].nunique(),
    df['Search Name'].nunique() - df_hb['Search Name'].nunique(),
    '',
    'TOBBWI & TOBBJN (Universal)',
    df[df['Rate Code'] == 'TOBBWI']['Room Nights'].sum(),
    df[df['Rate Code'] == 'TOBBJN']['Room Nights'].sum(),
])

exec_summary['Notes'].extend([
    'All bookings in dataset',
    'Total room nights across all bookings',
    'Total revenue generated',
    'Average rate per room night',
    '',
    '14% of all bookings',
    'Only 2.7% of total room nights!',
    'Only 3.1% of total revenue',
    'AED 26 premium vs overall average',
    '',
    'CRITICAL: 86% of bookings have NO F&B',
    'Massive opportunity gap',
    'Low share indicates upsell potential',
    '',
    'Total travel agencies in dataset',
    'Have at least 1 HB booking',
    'ZERO HB bookings - untapped!',
    '',
    'Serve all markets - prime for bundling',
    '2,915 nights (low HB attachment)',
    '2,530 nights (low HB attachment)',
])

df_exec = pd.DataFrame(exec_summary)

# ============================================================================
# SHEET 2: AGENCY DEEP DIVE - TOP 20
# ============================================================================
print("[2/10] Building Agency Deep Dive Analysis...")

# Calculate comprehensive agency metrics
agency_analysis = []

for agency in df['Search Name'].dropna().unique():
    agency_data = df[df['Search Name'] == agency]
    agency_hb = agency_data[agency_data['Has_HB']]

    total_nights = agency_data['Room Nights'].sum()
    hb_nights = agency_hb['Room Nights'].sum()

    analysis = {
        'Agency Name': agency,
        'Total Room Nights': total_nights,
        'HB Room Nights': hb_nights,
        'Non-HB Room Nights': total_nights - hb_nights,
        '% HB Nights': (hb_nights / total_nights * 100) if total_nights > 0 else 0,
        'Total Revenue (AED)': agency_data['Room Revenue'].sum(),
        'HB Revenue (AED)': agency_hb['Room Revenue'].sum(),
        'Non-HB Revenue (AED)': agency_data[~agency_data['Has_HB']]['Room Revenue'].sum(),
        '% HB Revenue': (agency_hb['Room Revenue'].sum() / agency_data['Room Revenue'].sum() * 100) if agency_data['Room Revenue'].sum() > 0 else 0,
        'Total Bookings': len(agency_data),
        'HB Bookings': len(agency_hb),
        'Avg Rate (Overall) AED': agency_data['Avg_Rate_Per_Night'].mean(),
        'Avg Rate (HB) AED': agency_hb['Avg_Rate_Per_Night'].mean() if len(agency_hb) > 0 else 0,
        'Avg Rate (Non-HB) AED': agency_data[~agency_data['Has_HB']]['Avg_Rate_Per_Night'].mean() if len(agency_data[~agency_data['Has_HB']]) > 0 else 0,
        'Top Rate Code': agency_data['Rate Code'].mode()[0] if len(agency_data['Rate Code'].mode()) > 0 else 'N/A',
        'Avg Nights per Booking': agency_data['Room Nights'].mean(),
    }
    agency_analysis.append(analysis)

df_agency = pd.DataFrame(agency_analysis)
df_agency = df_agency.sort_values('Total Revenue (AED)', ascending=False).reset_index(drop=True)
df_agency.index = df_agency.index + 1  # Start ranking from 1

# Top 20 for main sheet
df_agency_top20 = df_agency.head(20).copy()

print(f"  - Analyzed {len(df_agency)} agencies")
print(f"  - Top agency: {df_agency.iloc[0]['Agency Name']}")

# ============================================================================
# SHEET 3: MIRACLE TOURISM DEEP DIVE
# ============================================================================
print("[3/10] Building Miracle Tourism Analysis...")

miracle_data = df[df['Search Name'] == 'MIRACLE TOURISM LLC']
miracle_hb = miracle_data[miracle_data['Has_HB']]

# Miracle rate code analysis
miracle_rate_analysis = miracle_data.groupby(['Rate Code', 'Has_HB']).agg({
    'Room Nights': 'sum',
    'Room Revenue': 'sum',
    'Avg_Rate_Per_Night': 'mean'
}).round(2).reset_index()
miracle_rate_analysis['HB Status'] = miracle_rate_analysis['Has_HB'].map({True: 'With HB', False: 'Without HB'})
miracle_rate_analysis = miracle_rate_analysis[['Rate Code', 'HB Status', 'Room Nights', 'Room Revenue', 'Avg_Rate_Per_Night']]
miracle_rate_analysis.columns = ['Rate Code', 'HB Status', 'Room Nights', 'Room Revenue (AED)', 'Avg Rate (AED)']

# Booking size analysis
miracle_data['Booking_Size'] = pd.cut(miracle_data['Room Nights'],
                                       bins=[0, 10, 30, 50, 100, 1000],
                                       labels=['1-10 nights', '11-30 nights', '31-50 nights', '51-100 nights', '100+ nights'])

miracle_booking_analysis = miracle_data.groupby('Booking_Size', observed=True).agg({
    'Room Nights': ['count', 'sum'],
    'Room Revenue': 'sum',
    'Has_HB': 'sum'
}).round(2)

miracle_booking_analysis.columns = ['Bookings', 'Total Room Nights', 'Revenue (AED)', 'HB Bookings']
miracle_booking_analysis['% of Bookings'] = (miracle_booking_analysis['Bookings'] / len(miracle_data) * 100).round(1)
miracle_booking_analysis['% HB'] = (miracle_booking_analysis['HB Bookings'] / miracle_booking_analysis['Bookings'] * 100).round(1)
miracle_booking_analysis['Avg Revenue per Booking'] = (miracle_booking_analysis['Revenue (AED)'] / miracle_booking_analysis['Bookings']).round(2)
miracle_booking_analysis = miracle_booking_analysis.reset_index()

# Miracle summary stats
miracle_summary = pd.DataFrame({
    'Metric': [
        'Total Bookings',
        'Total Room Nights',
        'Total Revenue (AED)',
        'Avg Nights per Booking',
        '',
        'HB Bookings',
        'HB Room Nights',
        'HB Revenue (AED)',
        'HB Penetration Rate (%)',
        'HB Avg Nights per Booking',
        '',
        'Market',
        'Primary Rate Code',
        'Average Rate (Overall)',
        'Average Rate (HB)',
        '',
        'WHY #1 in HB:',
        '1. Booking Size',
        '2. Consistent HB Adoption',
        '3. Luxembourg Market Profile',
        '4. Rate Code Strategy',
    ],
    'Value': [
        len(miracle_data),
        miracle_data['Room Nights'].sum(),
        f"AED {miracle_data['Room Revenue'].sum():,.2f}",
        f"{miracle_data['Room Nights'].mean():.1f} nights",
        '',
        len(miracle_hb),
        miracle_hb['Room Nights'].sum(),
        f"AED {miracle_hb['Room Revenue'].sum():,.2f}",
        f"{len(miracle_hb)/len(miracle_data)*100:.1f}%",
        f"{miracle_hb['Room Nights'].mean():.1f} nights",
        '',
        'Luxembourg (TOMILUX rate code)',
        miracle_data['Rate Code'].mode()[0] if len(miracle_data) > 0 else 'N/A',
        f"AED {miracle_data['Avg_Rate_Per_Night'].mean():.2f}",
        f"AED {miracle_hb['Avg_Rate_Per_Night'].mean():.2f}",
        '',
        '',
        f"Avg {miracle_data['Room Nights'].mean():.1f} nights per booking - LARGE bookings",
        f"{len(miracle_hb)/len(miracle_data)*100:.0f}% of bookings include HB",
        'Luxembourg guests prefer full-package holidays',
        'TOMILUX code optimized for their market',
    ]
})

print(f"  - Miracle Tourism: {len(miracle_data)} bookings, {miracle_data['Room Nights'].mean():.1f} avg nights per booking")

# ============================================================================
# SHEET 4: UNIVERSAL RATE CODES (TOBBWI & TOBBJN)
# ============================================================================
print("[4/10] Analyzing Universal Rate Codes...")

universal_codes = ['TOBBWI', 'TOBBJN']
universal_analysis = []

for code in universal_codes:
    code_data = df[df['Rate Code'] == code]
    code_hb = code_data[code_data['Has_HB']]

    analysis = {
        'Rate Code': code,
        'Description': 'Universal - All Markets',
        'Total Bookings': len(code_data),
        'Total Room Nights': code_data['Room Nights'].sum(),
        'Total Revenue (AED)': code_data['Room Revenue'].sum(),
        'HB Bookings': len(code_hb),
        'HB Room Nights': code_hb['Room Nights'].sum(),
        'HB Revenue (AED)': code_hb['Room Revenue'].sum(),
        '% HB Penetration': (len(code_hb) / len(code_data) * 100) if len(code_data) > 0 else 0,
        'Avg Nights per Booking': code_data['Room Nights'].mean(),
        'Avg Nights per Booking (HB)': code_hb['Room Nights'].mean() if len(code_hb) > 0 else 0,
        'Avg Rate (Overall) AED': code_data['Avg_Rate_Per_Night'].mean(),
        'Avg Rate (HB) AED': code_hb['Avg_Rate_Per_Night'].mean() if len(code_hb) > 0 else 0,
        'Number of Agencies Using': code_data['Search Name'].nunique(),
        'Top Agency': code_data.groupby('Search Name')['Room Revenue'].sum().idxmax(),
    }
    universal_analysis.append(analysis)

# Agency breakdown for each universal code
tobbwi_agencies = df[df['Rate Code'] == 'TOBBWI'].groupby('Search Name').agg({
    'Room Nights': 'sum',
    'Room Revenue': 'sum',
    'Has_HB': 'sum'
}).round(2).reset_index()
tobbwi_agencies.columns = ['Agency', 'Room Nights', 'Revenue (AED)', 'HB Bookings']
tobbwi_agencies['% HB'] = (tobbwi_agencies['HB Bookings'] / df[df['Rate Code'] == 'TOBBWI'].groupby('Search Name').size().values * 100).round(1)
tobbwi_agencies = tobbwi_agencies.sort_values('Revenue (AED)', ascending=False).head(15)

tobbjn_agencies = df[df['Rate Code'] == 'TOBBJN'].groupby('Search Name').agg({
    'Room Nights': 'sum',
    'Room Revenue': 'sum',
    'Has_HB': 'sum'
}).round(2).reset_index()
tobbjn_agencies.columns = ['Agency', 'Room Nights', 'Revenue (AED)', 'HB Bookings']
tobbjn_agencies['% HB'] = (tobbjn_agencies['HB Bookings'] / df[df['Rate Code'] == 'TOBBJN'].groupby('Search Name').size().values * 100).round(1)
tobbjn_agencies = tobbjn_agencies.sort_values('Revenue (AED)', ascending=False).head(15)

df_universal = pd.DataFrame(universal_analysis)

print(f"  - TOBBWI: {len(df[df['Rate Code'] == 'TOBBWI'])} bookings")
print(f"  - TOBBJN: {len(df[df['Rate Code'] == 'TOBBJN'])} bookings")

# ============================================================================
# SHEET 5: HB CONVERSION OPPORTUNITY MATRIX
# ============================================================================
print("[5/10] Building Opportunity Matrix...")

opportunity_matrix = []

for agency in df['Search Name'].dropna().unique():
    agency_data = df[df['Search Name'] == agency]
    agency_hb = agency_data[agency_data['Has_HB']]

    total_nights = agency_data['Room Nights'].sum()
    hb_nights = agency_hb['Room Nights'].sum()
    current_hb_pct = (hb_nights / total_nights * 100) if total_nights > 0 else 0

    # Calculate potential if HB increased to 35%
    potential_hb_nights = total_nights * 0.35
    incremental_nights = potential_hb_nights - hb_nights

    # Estimate incremental revenue (assuming AED 120 per HB night)
    incremental_revenue = incremental_nights * 120 if incremental_nights > 0 else 0

    # Priority scoring
    volume_score = min(total_nights / 100, 10)  # Max 10 points for volume
    opportunity_score = (100 - current_hb_pct) / 10  # Max 10 points for low HB%
    priority_score = (volume_score + opportunity_score) / 2

    opportunity = {
        'Agency Name': agency,
        'Total Room Nights': total_nights,
        'Current HB Nights': hb_nights,
        'Current HB %': current_hb_pct,
        'Target HB % (35%)': 35.0,
        'Gap to Target (%)': max(35 - current_hb_pct, 0),
        'Potential HB Nights': max(incremental_nights, 0),
        'Est. Incremental F&B Revenue (AED)': max(incremental_revenue, 0),
        'Volume Score (1-10)': min(volume_score, 10),
        'Opportunity Score (1-10)': min(opportunity_score, 10),
        'Priority Score': priority_score,
        'Action Priority': 'HIGH' if priority_score >= 7 else ('MEDIUM' if priority_score >= 4 else 'LOW'),
        'Avg Nights per Booking': agency_data['Room Nights'].mean(),
        'Recommended Tactic': ''
    }

    # Assign tactics based on profile
    if total_nights > 500 and current_hb_pct < 10:
        opportunity['Recommended Tactic'] = 'Urgent: Executive meeting + Commission incentive'
    elif total_nights > 300 and current_hb_pct < 20:
        opportunity['Recommended Tactic'] = 'High Priority: Targeted HB promotion + Training'
    elif current_hb_pct > 50:
        opportunity['Recommended Tactic'] = 'Optimize: Upsell to premium HB packages'
    elif agency_data['Room Nights'].mean() > 50:
        opportunity['Recommended Tactic'] = 'Large Booking Leverage: Bundle HB in packages'
    else:
        opportunity['Recommended Tactic'] = 'Standard: Include in HB marketing campaign'

    opportunity_matrix.append(opportunity)

df_opportunity = pd.DataFrame(opportunity_matrix)
df_opportunity = df_opportunity.sort_values('Priority Score', ascending=False).reset_index(drop=True)
df_opportunity.index = df_opportunity.index + 1

print(f"  - Identified {len(df_opportunity[df_opportunity['Action Priority'] == 'HIGH'])} HIGH priority opportunities")
print(f"  - Total potential incremental revenue: AED {df_opportunity['Est. Incremental F&B Revenue (AED)'].sum():,.0f}")

# ============================================================================
# SHEET 6: MARKET SEGMENTATION
# ============================================================================
print("[6/10] Performing Market Segmentation...")

# Identify markets from rate codes
def identify_market(rate_code):
    if pd.isna(rate_code):
        return 'Unknown'
    rate_code = str(rate_code).upper()
    if 'CIS' in rate_code:
        return 'CIS Markets'
    elif 'MILUX' in rate_code:
        return 'Luxembourg'
    elif 'BBWI' in rate_code or 'BBJN' in rate_code or 'BB-WI' in rate_code or rate_code == 'TOBB':
        return 'Universal/Multi-Market'
    elif 'ROWI' in rate_code:
        return 'Specific Market'
    elif 'SSE' in rate_code or 'FSSE' in rate_code:
        return 'Secret Escapes'
    elif 'DG' in rate_code:
        return 'Desert Gate'
    elif 'EX' in rate_code:
        return 'Express/Quick'
    else:
        return 'Other'

df['Market_Segment'] = df['Rate Code'].apply(identify_market)

market_analysis = df.groupby('Market_Segment').agg({
    'Room Nights': ['count', 'sum', 'mean'],
    'Room Revenue': 'sum',
    'Has_HB': ['sum', 'mean'],
    'Avg_Rate_Per_Night': 'mean'
}).round(2)

market_analysis.columns = ['Bookings', 'Total Room Nights', 'Avg Nights per Booking', 'Total Revenue (AED)', 'HB Bookings', 'HB Penetration', 'Avg Rate (AED)']
market_analysis['HB Penetration'] = (market_analysis['HB Penetration'] * 100).round(1)
market_analysis['% of Total Revenue'] = (market_analysis['Total Revenue (AED)'] / df['Room Revenue'].sum() * 100).round(1)
market_analysis = market_analysis.sort_values('Total Revenue (AED)', ascending=False).reset_index()

# Top agencies by market
market_agency_detail = []
for market in market_analysis['Market_Segment'].unique():
    market_data = df[df['Market_Segment'] == market]
    top_agencies = market_data.groupby('Search Name').agg({
        'Room Nights': 'sum',
        'Room Revenue': 'sum',
        'Has_HB': 'sum'
    }).reset_index()
    top_agencies.columns = ['Agency', 'Room Nights', 'Revenue (AED)', 'HB Bookings']
    top_agencies['Market'] = market
    top_agencies = top_agencies.sort_values('Revenue (AED)', ascending=False).head(5)
    market_agency_detail.append(top_agencies[['Market', 'Agency', 'Room Nights', 'Revenue (AED)', 'HB Bookings']])

df_market_agencies = pd.concat(market_agency_detail, ignore_index=True)

print(f"  - Identified {len(market_analysis)} market segments")

# ============================================================================
# SHEET 7: PRESCRIPTIVE ACTION PLAN
# ============================================================================
print("[7/10] Creating Prescriptive Action Plan...")

action_plan = []

# Priority 1: High volume, low HB agencies
high_priority_agencies = df_opportunity[
    (df_opportunity['Total Room Nights'] > 300) &
    (df_opportunity['Current HB %'] < 15)
].head(5)

for idx, row in high_priority_agencies.iterrows():
    action_plan.append({
        'Priority': 1,
        'Category': 'High Volume - Low HB Conversion',
        'Target': row['Agency Name'],
        'Current State': f"{row['Total Room Nights']:.0f} nights, {row['Current HB %']:.1f}% HB",
        'Opportunity': f"Potential {row['Potential HB Nights']:.0f} additional HB nights",
        'Action Required': 'Executive Meeting + Commission Incentive (15% on HB bookings)',
        'Timeline': 'Week 1-2',
        'Est. Impact (AED)': f"{row['Est. Incremental F&B Revenue (AED)']:,.0f}",
        'Success Metric': 'Achieve 35% HB penetration within 60 days',
        'Owner': 'Sales Director'
    })

# Priority 2: Universal rate code optimization
action_plan.append({
    'Priority': 2,
    'Category': 'Rate Code Optimization',
    'Target': 'TOBBWI & TOBBJN (Universal Codes)',
    'Current State': f"2,915 + 2,530 = 5,445 room nights, Low HB attachment",
    'Opportunity': 'These codes serve all markets - prime for HB bundling',
    'Action Required': 'Create HB variants: TOBBWI-HB, TOBBJN-HB with +AED 120 premium',
    'Timeline': 'Week 2-3',
    'Est. Impact (AED)': '200,000+',
    'Success Metric': '40% of bookings use HB variant',
    'Owner': 'Revenue Manager'
})

# Priority 3: Replicate Miracle success (Luxembourg model)
action_plan.append({
    'Priority': 3,
    'Category': 'Best Practice Replication',
    'Target': 'Luxembourg Market Model (Miracle Tourism)',
    'Current State': 'Miracle has strong HB adoption with large bookings',
    'Opportunity': 'Apply same model to other markets with large bookings',
    'Action Required': 'Identify agencies with large bookings, offer package deals',
    'Timeline': 'Week 3-4',
    'Est. Impact (AED)': '150,000+',
    'Success Metric': 'Identify 3 new agencies with similar booking patterns',
    'Owner': 'Business Development'
})

# Priority 4: CIS market expansion
action_plan.append({
    'Priority': 4,
    'Category': 'CIS Market Expansion',
    'Target': 'CIS Markets (TOKHACIS, TOVOYCIS, TOCIS25)',
    'Current State': '36% HB penetration (best in dataset), only 9 HB bookings total',
    'Opportunity': 'CIS guests WANT packages - scale up volume',
    'Action Required': 'Launch CIS agency blitz: Target 10 new Russian/Eastern European agencies',
    'Timeline': 'Week 1-4',
    'Est. Impact (AED)': '80,000+',
    'Success Metric': '50 CIS HB bookings in 90 days',
    'Owner': 'International Sales'
})

# Priority 5: Large booking strategy
action_plan.append({
    'Priority': 5,
    'Category': 'Large Booking Strategy',
    'Target': 'Bookings with 50+ nights',
    'Current State': f"{len(df[df['Room Nights'] >= 50])} bookings ≥50 nights, mixed HB adoption",
    'Opportunity': 'Large bookings = higher F&B consumption opportunity',
    'Action Required': 'Implement policy: 50+ night bookings = HB mandatory or heavily discounted',
    'Timeline': 'Week 4-5',
    'Est. Impact (AED)': '120,000+',
    'Success Metric': '80% of 50+ night bookings include HB',
    'Owner': 'Reservations Manager'
})

# Additional tactical actions
additional_actions = [
    {
        'Priority': 6,
        'Category': 'Training & Enablement',
        'Target': 'Top 20 agencies without HB',
        'Current State': 'Agencies may not understand HB value proposition',
        'Opportunity': 'Education = Conversion',
        'Action Required': 'Webinar series: "Why HB Increases Guest Satisfaction" + Sales toolkit',
        'Timeline': 'Week 2-6',
        'Est. Impact (AED)': '50,000',
        'Success Metric': '80% attendance, 25% conversion',
        'Owner': 'Training Manager'
    },
    {
        'Priority': 7,
        'Category': 'Product Innovation',
        'Target': 'All Markets',
        'Current State': 'Only basic HB offered',
        'Opportunity': 'Tiered packages increase average transaction',
        'Action Required': 'Launch 3 tiers: Essential HB (AED 120), Enhanced HB (AED 160), Premium HB (AED 200)',
        'Timeline': 'Week 5-8',
        'Est. Impact (AED)': '75,000',
        'Success Metric': '20% opt for Enhanced/Premium',
        'Owner': 'F&B Director'
    },
    {
        'Priority': 8,
        'Category': 'Agent Incentive Program',
        'Target': 'All agencies',
        'Current State': 'No specific HB incentives',
        'Opportunity': 'Financial motivation drives behavior',
        'Action Required': 'Tiered commission: Book 50 HB nights = +2%, 100 nights = +3%, 200 nights = +5%',
        'Timeline': 'Week 1 (Launch immediately)',
        'Est. Impact (AED)': '100,000+',
        'Success Metric': '15 agencies qualify for bonus tiers',
        'Owner': 'Sales Director'
    },
    {
        'Priority': 9,
        'Category': 'Data & Tracking',
        'Target': 'Internal Systems',
        'Current State': 'No real-time HB tracking dashboard',
        'Opportunity': 'What gets measured gets managed',
        'Action Required': 'Build weekly HB dashboard: Track by agency, rate code, market',
        'Timeline': 'Week 3-4',
        'Est. Impact (AED)': 'Enabler for all other initiatives',
        'Success Metric': 'Dashboard live, weekly reviews',
        'Owner': 'Revenue Analyst'
    },
    {
        'Priority': 10,
        'Category': 'Competitive Analysis',
        'Target': 'Market Positioning',
        'Current State': 'Unknown how our HB offering compares',
        'Opportunity': 'Ensure competitive pricing and offering',
        'Action Required': 'Shop 5 competitors: Compare HB pricing, inclusions, guest feedback',
        'Timeline': 'Week 2-3',
        'Est. Impact (AED)': 'Strategic input',
        'Success Metric': 'Competitive report completed',
        'Owner': 'Strategy Manager'
    }
]

action_plan.extend(additional_actions)
df_action_plan = pd.DataFrame(action_plan)

print(f"  - Created {len(action_plan)} prioritized action items")

# ============================================================================
# SHEET 8: EDA VISUALIZATION DATA
# ============================================================================
print("[8/10] Preparing EDA Visualization Data...")

# 1. Top 15 Agencies Comparison (Total vs HB)
top15_agencies_comp = df_agency.head(15)[['Agency Name', 'Total Room Nights', 'HB Room Nights', 'Total Revenue (AED)', 'HB Revenue (AED)']].copy()

# 2. Top 15 Agencies by HB Performance
top15_hb_performance = df_agency[df_agency['HB Room Nights'] > 0].sort_values('HB Room Nights', ascending=False).head(15)[
    ['Agency Name', 'HB Room Nights', 'HB Revenue (AED)', '% HB Nights']
].copy()

# 3. Rate Code Performance Comparison
rate_code_performance = df.groupby('Rate Code').agg({
    'Room Nights': 'sum',
    'Room Revenue': 'sum',
    'Has_HB': 'sum'
}).reset_index()
rate_code_performance.columns = ['Rate Code', 'Total Room Nights', 'Total Revenue (AED)', 'HB Bookings']
rate_code_performance = rate_code_performance.sort_values('Total Revenue (AED)', ascending=False).head(15)

# 4. HB vs Non-HB Comparison
hb_comparison = pd.DataFrame({
    'Category': ['Half Board', 'Non-Half Board'],
    'Bookings': [len(df_hb), len(df_non_hb)],
    'Room Nights': [df_hb['Room Nights'].sum(), df_non_hb['Room Nights'].sum()],
    'Revenue (AED)': [df_hb['Room Revenue'].sum(), df_non_hb['Room Revenue'].sum()],
    'Avg Rate (AED)': [df_hb['Avg_Rate_Per_Night'].mean(), df_non_hb['Avg_Rate_Per_Night'].mean()]
})

# 5. Market Segment Comparison
market_segment_comp = market_analysis[['Market_Segment', 'Total Room Nights', 'Total Revenue (AED)', 'HB Bookings', 'HB Penetration']].copy()

# 6. Top 10 CIS Performance
df_cis = df[df['Rate Code'].str.contains('CIS', case=False, na=False)]
if len(df_cis) > 0:
    cis_performance = df_cis.groupby('Search Name').agg({
        'Room Nights': 'sum',
        'Room Revenue': 'sum',
        'Has_HB': 'sum'
    }).reset_index()
    cis_performance.columns = ['Agency', 'Room Nights', 'Revenue (AED)', 'HB Bookings']
    cis_performance = cis_performance.sort_values('Revenue (AED)', ascending=False).head(10)
else:
    cis_performance = pd.DataFrame()

# 7. Booking Size Distribution
df['Booking_Size_Category'] = pd.cut(df['Room Nights'],
                                      bins=[0, 5, 15, 30, 50, 1000],
                                      labels=['1-5 nights', '6-15 nights', '16-30 nights', '31-50 nights', '50+ nights'])
booking_size_dist = df.groupby('Booking_Size_Category', observed=True).agg({
    'Room Nights': 'count',
    'Has_HB': 'sum'
}).reset_index()
booking_size_dist.columns = ['Booking Size', 'Count', 'HB Bookings']
booking_size_dist['HB %'] = (booking_size_dist['HB Bookings'] / booking_size_dist['Count'] * 100).round(1)

# ============================================================================
# SHEET 9: RAW DATA REFERENCE
# ============================================================================
print("[9/10] Preparing raw data reference...")

df_raw = df.copy()
df_raw = df_raw.sort_values('Room Revenue', ascending=False)

# ============================================================================
# CREATE EXCEL WORKBOOK WITH FORMATTING AND CHARTS
# ============================================================================
print("\n[10/10] Creating Formatted Excel Workbook with Charts...")
print("="*80)

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=10)
highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
high_priority_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
medium_priority_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
low_priority_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_sheet_header(ws, end_col):
    """Apply header formatting to first row"""
    for col in range(1, end_col + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def format_currency_column(ws, col_idx, start_row=2):
    """Format column as currency"""
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row, col_idx)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

def format_percentage_column(ws, col_idx, start_row=2):
    """Format column as percentage"""
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row, col_idx)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.0%'
            if cell.value > 1:
                cell.value = cell.value / 100

def auto_size_columns(ws):
    """Auto-size all columns"""
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Create all sheets
print("Creating sheets...")

# Sheet 1: Executive Summary
ws1 = wb.create_sheet("Executive Summary")
for r in dataframe_to_rows(df_exec, index=False, header=True):
    ws1.append(r)
ws1.insert_rows(1)
ws1['A1'] = 'HALF BOARD ANALYSIS - EXECUTIVE SUMMARY'
ws1['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws1.merge_cells('A1:C1')
format_sheet_header(ws1, 3)
auto_size_columns(ws1)
for row in [7, 11, 14, 18]:
    ws1[f'A{row}'].fill = highlight_fill
    ws1[f'B{row}'].fill = highlight_fill
    ws1[f'C{row}'].fill = highlight_fill

# Sheet 2: Agency Deep Dive
ws2 = wb.create_sheet("Agency Deep Dive - Top 20")
for r in dataframe_to_rows(df_agency_top20, index=True, header=True):
    ws2.append(r)
format_sheet_header(ws2, len(df_agency_top20.columns) + 1)
format_currency_column(ws2, 6)
format_currency_column(ws2, 7)
format_currency_column(ws2, 8)
format_percentage_column(ws2, 5)
format_percentage_column(ws2, 9)
auto_size_columns(ws2)
for row in [2, 3, 4]:
    for col in range(1, len(df_agency_top20.columns) + 2):
        ws2.cell(row, col).fill = highlight_fill

# Sheet 3: Miracle Tourism
ws3 = wb.create_sheet("Miracle Tourism Deep Dive")
ws3['A1'] = 'MIRACLE TOURISM LLC - DETAILED ANALYSIS (Luxembourg Market)'
ws3['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws3.merge_cells('A1:D1')
ws3['A3'] = 'SUMMARY STATISTICS'
ws3['A3'].font = Font(bold=True, size=12)
start_row = 4
for r in dataframe_to_rows(miracle_summary, index=False, header=True):
    ws3.append(r)
ws3[f'A{start_row + 22}'] = 'BOOKING SIZE DISTRIBUTION'
ws3[f'A{start_row + 22}'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(miracle_booking_analysis, index=False, header=True):
    ws3.append(r)
ws3[f'A{ws3.max_row + 2}'] = 'RATE CODE BREAKDOWN'
ws3[f'A{ws3.max_row}'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(miracle_rate_analysis, index=False, header=True):
    ws3.append(r)
auto_size_columns(ws3)

# Sheet 4: Universal Rate Codes
ws4 = wb.create_sheet("Universal Rate Codes")
ws4['A1'] = 'UNIVERSAL RATE CODES ANALYSIS (TOBBWI & TOBBJN)'
ws4['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws4.merge_cells('A1:E1')
ws4['A3'] = 'RATE CODE SUMMARY'
ws4['A3'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(df_universal, index=False, header=True):
    ws4.append(r)
ws4[f'A{ws4.max_row + 2}'] = 'TOP AGENCIES - TOBBWI'
ws4[f'A{ws4.max_row}'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(tobbwi_agencies, index=False, header=True):
    ws4.append(r)
ws4[f'A{ws4.max_row + 2}'] = 'TOP AGENCIES - TOBBJN'
ws4[f'A{ws4.max_row}'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(tobbjn_agencies, index=False, header=True):
    ws4.append(r)
auto_size_columns(ws4)

# Sheet 5: Opportunity Matrix
ws5 = wb.create_sheet("Opportunity Matrix")
for r in dataframe_to_rows(df_opportunity.head(30), index=True, header=True):
    ws5.append(r)
format_sheet_header(ws5, len(df_opportunity.columns) + 1)
format_currency_column(ws5, 8)
format_percentage_column(ws5, 4)
format_percentage_column(ws5, 6)
auto_size_columns(ws5)
for row in range(2, ws5.max_row + 1):
    priority_cell = ws5.cell(row, 13)
    if priority_cell.value == 'HIGH':
        for col in range(1, ws5.max_column + 1):
            ws5.cell(row, col).fill = high_priority_fill
    elif priority_cell.value == 'MEDIUM':
        for col in range(1, ws5.max_column + 1):
            ws5.cell(row, col).fill = medium_priority_fill
    elif priority_cell.value == 'LOW':
        for col in range(1, ws5.max_column + 1):
            ws5.cell(row, col).fill = low_priority_fill

# Sheet 6: Market Segmentation
ws6 = wb.create_sheet("Market Segmentation")
ws6['A1'] = 'MARKET SEGMENTATION ANALYSIS'
ws6['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws6.merge_cells('A1:E1')
ws6['A3'] = 'MARKET PERFORMANCE SUMMARY'
ws6['A3'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(market_analysis, index=False, header=True):
    ws6.append(r)
ws6[f'A{ws6.max_row + 2}'] = 'TOP AGENCIES BY MARKET'
ws6[f'A{ws6.max_row}'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(df_market_agencies, index=False, header=True):
    ws6.append(r)
auto_size_columns(ws6)

# Sheet 7: Action Plan
ws7 = wb.create_sheet("ACTION PLAN")
for r in dataframe_to_rows(df_action_plan, index=False, header=True):
    ws7.append(r)
format_sheet_header(ws7, len(df_action_plan.columns))
auto_size_columns(ws7)
for row in range(2, ws7.max_row + 1):
    priority_cell = ws7.cell(row, 1)
    if priority_cell.value <= 3:
        for col in range(1, ws7.max_column + 1):
            ws7.cell(row, col).fill = high_priority_fill
    elif priority_cell.value <= 7:
        for col in range(1, ws7.max_column + 1):
            ws7.cell(row, col).fill = medium_priority_fill

# Sheet 8: EDA VISUALIZATIONS
print("Creating EDA Visualization sheet with charts...")
ws8 = wb.create_sheet("EDA VISUALIZATIONS")
ws8['A1'] = 'EXPLORATORY DATA ANALYSIS - VISUAL COMPARISONS'
ws8['A1'].font = Font(bold=True, size=16, color="1F4E78")
ws8.merge_cells('A1:H1')

current_row = 3

# Table 1: HB vs Non-HB Comparison
ws8[f'A{current_row}'] = 'TABLE 1: HALF BOARD vs NON-HALF BOARD COMPARISON'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(hb_comparison, index=False, header=True):
    ws8.append(r)
current_row = ws8.max_row

# Chart 1: HB vs Non-HB Pie Chart
chart1 = PieChart()
chart1.title = "Revenue Distribution: HB vs Non-HB"
data = Reference(ws8, min_col=4, min_row=current_row-2, max_row=current_row)
labels = Reference(ws8, min_col=1, min_row=current_row-1, max_row=current_row)
chart1.add_data(data, titles_from_data=False)
chart1.set_categories(labels)
chart1.height = 10
chart1.width = 15
ws8.add_chart(chart1, f'F{current_row-3}')

current_row += 3

# Table 2: Top 15 Agencies - Total vs HB
ws8[f'A{current_row}'] = 'TABLE 2: TOP 15 AGENCIES - TOTAL vs HALF BOARD PERFORMANCE'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(top15_agencies_comp, index=False, header=True):
    ws8.append(r)
chart_start = current_row + 1
current_row = ws8.max_row

# Chart 2: Top 15 Agencies Bar Chart
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Top 15 Agencies: Total Room Nights vs HB Room Nights"
chart2.y_axis.title = 'Room Nights'
chart2.x_axis.title = 'Agency'
data = Reference(ws8, min_col=2, min_row=chart_start, max_row=current_row, max_col=3)
cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.height = 12
chart2.width = 20
ws8.add_chart(chart2, f'F{chart_start}')

current_row += 3

# Table 3: Top 15 HB Performers
ws8[f'A{current_row}'] = 'TABLE 3: TOP 15 HALF BOARD PERFORMERS'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(top15_hb_performance, index=False, header=True):
    ws8.append(r)
chart_start = current_row + 1
current_row = ws8.max_row

# Chart 3: Top HB Performers
chart3 = BarChart()
chart3.type = "bar"
chart3.title = "Top 15 HB Performers by Room Nights"
chart3.x_axis.title = 'HB Room Nights'
chart3.y_axis.title = 'Agency'
data = Reference(ws8, min_col=2, min_row=chart_start, max_row=current_row, max_col=2)
cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
chart3.height = 15
chart3.width = 18
ws8.add_chart(chart3, f'F{chart_start}')

current_row += 3

# Table 4: Rate Code Performance
ws8[f'A{current_row}'] = 'TABLE 4: TOP 15 RATE CODES PERFORMANCE'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(rate_code_performance, index=False, header=True):
    ws8.append(r)
chart_start = current_row + 1
current_row = ws8.max_row

# Chart 4: Rate Code Performance
chart4 = BarChart()
chart4.type = "col"
chart4.title = "Top 15 Rate Codes by Revenue"
chart4.y_axis.title = 'Revenue (AED)'
chart4.x_axis.title = 'Rate Code'
data = Reference(ws8, min_col=3, min_row=chart_start, max_row=current_row, max_col=3)
cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
chart4.height = 12
chart4.width = 20
ws8.add_chart(chart4, f'F{chart_start}')

current_row += 3

# Table 5: Market Segment Performance
ws8[f'A{current_row}'] = 'TABLE 5: MARKET SEGMENT PERFORMANCE'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(market_segment_comp, index=False, header=True):
    ws8.append(r)
chart_start = current_row + 1
current_row = ws8.max_row

# Chart 5: Market Segment HB Penetration
chart5 = BarChart()
chart5.type = "col"
chart5.title = "HB Penetration by Market Segment"
chart5.y_axis.title = 'HB Penetration %'
chart5.x_axis.title = 'Market Segment'
data = Reference(ws8, min_col=5, min_row=chart_start, max_row=current_row, max_col=5)
cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
chart5.add_data(data, titles_from_data=True)
chart5.set_categories(cats)
chart5.height = 12
chart5.width = 18
ws8.add_chart(chart5, f'F{chart_start}')

current_row += 3

# Table 6: Booking Size Distribution
ws8[f'A{current_row}'] = 'TABLE 6: BOOKING SIZE DISTRIBUTION & HB PENETRATION'
ws8[f'A{current_row}'].font = Font(bold=True, size=12)
current_row += 1
for r in dataframe_to_rows(booking_size_dist, index=False, header=True):
    ws8.append(r)
chart_start = current_row + 1
current_row = ws8.max_row

# Chart 6: Booking Size vs HB %
chart6 = LineChart()
chart6.title = "HB Penetration by Booking Size"
chart6.y_axis.title = 'HB Penetration %'
chart6.x_axis.title = 'Booking Size'
data = Reference(ws8, min_col=4, min_row=chart_start, max_row=current_row, max_col=4)
cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
chart6.add_data(data, titles_from_data=True)
chart6.set_categories(cats)
chart6.height = 10
chart6.width = 15
ws8.add_chart(chart6, f'F{chart_start}')

# Table 7: CIS Performance (if available)
if len(cis_performance) > 0:
    current_row += 3
    ws8[f'A{current_row}'] = 'TABLE 7: TOP CIS MARKET AGENCIES'
    ws8[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1
    for r in dataframe_to_rows(cis_performance, index=False, header=True):
        ws8.append(r)
    chart_start = current_row + 1
    current_row = ws8.max_row

    # Chart 7: CIS Agencies
    chart7 = BarChart()
    chart7.type = "bar"
    chart7.title = "Top CIS Agencies by Revenue"
    chart7.x_axis.title = 'Revenue (AED)'
    chart7.y_axis.title = 'Agency'
    data = Reference(ws8, min_col=3, min_row=chart_start, max_row=current_row, max_col=3)
    cats = Reference(ws8, min_col=1, min_row=chart_start+1, max_row=current_row)
    chart7.add_data(data, titles_from_data=True)
    chart7.set_categories(cats)
    chart7.height = 12
    chart7.width = 15
    ws8.add_chart(chart7, f'F{chart_start}')

auto_size_columns(ws8)

# Sheet 9: Raw Data
ws9 = wb.create_sheet("Raw Data")
for r in dataframe_to_rows(df_raw, index=False, header=True):
    ws9.append(r)
format_sheet_header(ws9, len(df_raw.columns))
auto_size_columns(ws9)

# Save workbook
output_file = '/home/gee_devops254/Downloads/Half Board/Half_Board_Comprehensive_Analysis_Revised.xlsx'
wb.save(output_file)

print("\n" + "="*80)
print("✓ EXCEL WORKBOOK CREATED SUCCESSFULLY!")
print("="*80)
print(f"\nFile location: {output_file}")
print("\nWorkbook contains 9 sheets:")
print("  1. Executive Summary - Key metrics and insights")
print("  2. Agency Deep Dive - Top 20 agencies with HB breakdown")
print("  3. Miracle Tourism Deep Dive - Why they're #1 (Luxembourg market)")
print("  4. Universal Rate Codes - TOBBWI & TOBBJN analysis")
print("  5. Opportunity Matrix - Prioritized conversion opportunities")
print("  6. Market Segmentation - Performance by market type")
print("  7. ACTION PLAN - Prescriptive actions with priorities")
print("  8. EDA VISUALIZATIONS - Tables & Charts for comparison ⭐NEW")
print("  9. Raw Data - Complete dataset reference")
print("\n" + "="*80)
print("KEY CHANGES:")
print("="*80)
print("✓ Removed Average Length of Stay calculations")
print("✓ Added 'Avg Nights per Booking' (Room Nights / Bookings)")
print("✓ Created dedicated EDA VISUALIZATIONS sheet with:")
print("  - 7 comparison tables")
print("  - 7 interactive charts (Pie, Bar, Line)")
print("  - Visual analysis of HB performance across dimensions")
print("="*80)
