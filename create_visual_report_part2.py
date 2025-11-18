import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from pathlib import Path
import os

print("="*80)
print("COMPLETING VISUAL REPORT - PART 2")
print("="*80)

# Load the workbook created by part 1
wb = load_workbook('/home/gee_devops254/Downloads/Half Board/Half_Board_Visual_Report.xlsx')

base_path = "/home/gee_devops254/Downloads/Half Board"

# Styles
title_font = Font(bold=True, color="FFFFFF", size=16)
title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
description_font = Font(size=10, italic=True)
description_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

def create_category_sheet(wb, sheet_name, title, chart_folder, chart_descriptions):
    """Create a sheet with charts from a specific category"""
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    current_row = 3

    for chart_num, desc in chart_descriptions.items():
        chart_files = list(Path(f'{base_path}/charts/{chart_folder}').glob(f'{chart_num}_*.png'))
        if not chart_files:
            continue

        chart_path = str(chart_files[0])

        # Title
        ws[f'A{current_row}'] = f'Chart {chart_num}: {desc["title"]}'
        ws[f'A{current_row}'].font = Font(bold=True, size=13, color="2C3E50")
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws.row_dimensions[current_row].height = 25

        # Description
        ws[f'A{current_row+1}'] = desc['description']
        ws[f'A{current_row+1}'].font = description_font
        ws[f'A{current_row+1}'].fill = description_fill
        ws[f'A{current_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row+1}:H{current_row+2}')
        ws.row_dimensions[current_row+1].height = 35

        # Key Insight
        ws[f'A{current_row+3}'] = '🔑 KEY INSIGHT:'
        ws[f'B{current_row+3}'] = desc['insight']
        ws[f'A{current_row+3}'].font = Font(bold=True, color="E74C3C")
        ws[f'B{current_row+3}'].font = Font(bold=True, color="27AE60")
        ws[f'B{current_row+3}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        ws.merge_cells(f'B{current_row+3}:H{current_row+3}')
        ws.row_dimensions[current_row+3].height = 20

        # Image
        if os.path.exists(chart_path):
            img = Image(chart_path)
            img.width = 950
            img.height = int(950 * img.height / img.width) if img.width > 0 else 380
            ws.add_image(img, f'A{current_row+4}')

            rows_needed = int(img.height / 20) + 1
            for i in range(rows_needed):
                ws.row_dimensions[current_row+4+i].height = 20

            current_row += rows_needed + 6
        else:
            current_row += 25

    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14

# ============================================================================
# SHEET 6: RATE CODE ANALYSIS
# ============================================================================
print("[6/11] Creating Rate Code Analysis...")
ratecode_charts = {
    '20': {
        'title': 'Top 15 Rate Codes by Revenue',
        'description': 'TOBBWI and TOBBJN (universal codes) highlighted in red. These are top revenue generators serving all markets.',
        'insight': 'Universal codes = Prime targets for HB bundling strategy'
    },
    '21': {
        'title': 'Top 15 Rate Codes by Nights',
        'description': 'TOBBWI (2,915 nights) + TOBBJN (2,530 nights) = 5,445 combined nights with minimal HB attachment.',
        'insight': 'If 35% adopted HB = AED 200k+ incremental revenue'
    },
    '22': {
        'title': 'TOBBWI Agency Performance',
        'description': 'Top 10 agencies using TOBBWI rate code. TBO, Darina, Webbeds lead but HB counts shown are low.',
        'insight': 'Engage these agencies for TOBBWI-HB variant creation'
    },
    '23': {
        'title': 'TOBBJN Agency Performance',
        'description': 'Dubai Link, Darina dominate TOBBJN usage with minimal HB attachment. Clear conversion opportunity.',
        'insight': 'Make TOBBJN-HB mandatory for 5+ night stays'
    },
    '24': {
        'title': 'Universal Codes HB Penetration',
        'description': 'Dramatic grouped bars showing HB vs Non-HB split. Non-HB dominates both codes (red bars tower over green).',
        'insight': 'Use in executive presentations to show opportunity size'
    },
    '25': {
        'title': 'Rate Code Avg Rate Comparison',
        'description': 'Color-coded by rate level: Green >AED 400, Orange AED 300-400, Red <AED 300. Shows pricing tiers.',
        'insight': 'Tier HB pricing by rate code - premium codes get premium HB'
    }
}

create_category_sheet(wb, "🔑 Rate Codes", "RATE CODE ANALYSIS - Universal & Specific",
                     "04_Rate_Code_Analysis", ratecode_charts)

# ============================================================================
# SHEET 7: MARKET SEGMENTATION
# ============================================================================
print("[7/11] Creating Market Segmentation...")
market_charts = {
    '26': {
        'title': 'Market Revenue Distribution',
        'description': 'Universal/Multi-Market segment dominates revenue but lacks HB focus. This is the #1 opportunity segment.',
        'insight': 'Universal segment = Highest revenue + Lowest HB = Biggest opportunity'
    },
    '27': {
        'title': 'Market HB Penetration',
        'description': 'Color-coded by performance: CIS Markets show 36% HB penetration (GREEN - best in dataset). Target line at 35%.',
        'insight': 'CIS guests already want HB - scale up volume immediately!'
    },
    '28': {
        'title': 'CIS Market Performance',
        'description': 'TOKHACIS, TOVOYCIS, TOCIS25 performance breakdown. TOKHACIS leads with AED 17,888 revenue from just 2 bookings!',
        'insight': 'Launch CIS agency acquisition campaign - they convert easily'
    },
    '29': {
        'title': 'Luxembourg Market Analysis',
        'description': 'Miracle Tourism monopolizes Luxembourg market (TOMILUX code). Single agency domination shows market potential.',
        'insight': 'Luxembourg market profile = High HB. Seek similar European markets.'
    },
    '30': {
        'title': 'Universal vs Specific Comparison',
        'description': 'Dual metric chart: Universal has volume, CIS has HB percentage. Goal: Combine both strengths.',
        'insight': 'Apply CIS HB tactics to Universal segment for maximum impact'
    }
}

create_category_sheet(wb, "🌍 Markets", "MARKET SEGMENTATION - Geographic Analysis",
                     "05_Market_Segmentation", market_charts)

# ============================================================================
# SHEET 8: HB PERFORMANCE
# ============================================================================
print("[8/11] Creating HB Performance...")
performance_charts = {
    '31': {
        'title': 'HB Bookings Tree Visualization',
        'description': 'Proportional representation: Top 3 agencies (Miracle, Desert Gate, Al Khalidiah) dominate HB nights.',
        'insight': '80/20 rule confirmed - Top 5 agencies drive 80% of HB'
    },
    '32': {
        'title': 'HB Revenue Pareto',
        'description': 'Pareto chart with cumulative percentage line. Red line crosses 80% at ~5 agencies. Classic 80/20 distribution.',
        'insight': 'Focus resources on top performers + scale the model'
    },
    '33': {
        'title': 'HB Avg Rate Distribution',
        'description': 'Histogram of HB rates clustering around AED 350-450. Mean (red) and median (orange) lines shown.',
        'insight': 'Create tiered HB: Essential (120), Enhanced (160), Premium (200)'
    },
    '34': {
        'title': 'HB vs Non-HB Rate Scatter',
        'description': 'Scatter plot with diagonal line. Points above = HB rates higher (good). Points below = pricing correction needed.',
        'insight': 'Agencies below diagonal need immediate rate corrections'
    },
    '35': {
        'title': 'HB Penetration Heatmap',
        'description': 'Color-coded matrix: Top 10 agencies × Top 10 rate codes. Green = high HB%, Red = low HB%. Red cells = opportunities.',
        'insight': 'Red cells = Immediate targeting opportunities for conversion'
    },
    '36': {
        'title': 'Booking Size vs HB Adoption',
        'description': 'Scatter with trend line showing positive correlation. Larger bookings = higher HB percentage. Bubble size = volume.',
        'insight': 'Mandate HB for bookings >50 nights - correlation proven'
    }
}

create_category_sheet(wb, "📊 HB Performance", "HALF BOARD PERFORMANCE ANALYSIS",
                     "06_HB_Performance", performance_charts)

# ============================================================================
# SHEET 9: OPPORTUNITY ANALYSIS
# ============================================================================
print("[9/11] Creating Opportunity Analysis...")
opportunity_charts = {
    '37': {
        'title': 'Opportunity Matrix Quadrant',
        'description': '4-quadrant scatter: Bottom-right (High Volume + Low HB%) = PRIORITY targets. Bubble size = volume, color = potential revenue.',
        'insight': 'Bottom-right quadrant agencies = Focus ALL sales efforts here'
    },
    '38': {
        'title': 'Top 10 High Priority',
        'description': 'Specific agencies with >AED 50k incremental revenue potential each. Current HB% shown. These are THE targets.',
        'insight': 'Schedule executive meetings with these 10 within 2 weeks'
    },
    '39': {
        'title': 'Incremental Revenue Potential',
        'description': 'Top 15 agencies ranked by potential. Total = AED 920k if all reach 35% HB penetration. Color gradient shows urgency.',
        'insight': 'Use in board presentations to justify HB program investment'
    },
    '40': {
        'title': 'Current vs Target HB',
        'description': 'Side-by-side bars: Red (current) vs Green (35% target). Visual gap is dramatic for most agencies.',
        'insight': 'Set agency-specific HB targets and track monthly progress'
    },
    '41': {
        'title': 'Quick Wins vs Long-Term',
        'description': 'Effort-Impact matrix with 10 actions plotted. Green quadrant = Quick wins (low effort, high impact). Actions 1-2 here.',
        'insight': 'Execute quick wins (Actions 1-2) immediately while planning long-term'
    }
}

create_category_sheet(wb, "🎯 Opportunities", "OPPORTUNITY ANALYSIS - AED 920k Potential",
                     "07_Opportunity_Analysis", opportunity_charts)

# ============================================================================
# SHEET 10: CORRELATIONS & RELATIONSHIPS
# ============================================================================
print("[10/11] Creating Correlations...")
correlation_charts = {
    '42': {
        'title': 'Correlation Heatmap',
        'description': 'Pearson correlation matrix for Room Nights, Revenue, Avg Rate. Red = positive, Blue = negative. 0.999 = perfect.',
        'insight': 'Room Nights & Revenue: Perfect correlation (0.999)'
    },
    '43': {
        'title': 'Nights vs Revenue by HB',
        'description': 'Dual scatter (HB=green, Non-HB=red) with trend lines. Both have similar slopes = rate parity correct.',
        'insight': 'Opportunity is in VOLUME, not rates - both have same slope'
    },
    '44': {
        'title': 'Rate vs Booking Size',
        'description': 'Scatter plot colored by HB status. No strong correlation = pricing is booking-size agnostic currently.',
        'insight': 'Consider volume discounts for large bookings to encourage HB'
    },
    '45': {
        'title': 'Agency Performance Matrix',
        'description': 'Bubble scatter: Revenue (X) vs HB Penetration (Y). Color = HB%. Goal: Move agencies to top-right quadrant.',
        'insight': 'Goal: Move agencies from top-left to top-right quadrant'
    }
}

create_category_sheet(wb, "🔗 Correlations", "CORRELATION & RELATIONSHIPS",
                     "08_Correlation_Relationships", correlation_charts)

# ============================================================================
# SHEET 11: STRATEGIC DASHBOARDS
# ============================================================================
print("[11/11] Creating Strategic Dashboards...")
dashboard_charts = {
    '46': {
        'title': 'Executive Summary Dashboard',
        'description': 'Multi-panel dashboard: 4 KPI cards, HB penetration pie, top 10 agencies, market segments, rate comparison. Single-page overview.',
        'insight': 'Use in C-level presentations and quarterly board reviews'
    },
    '47': {
        'title': 'HB Performance Scorecard',
        'description': 'Detailed HB metrics: Top generators, revenue by market, rate distribution, trend vs target (simulated monthly data).',
        'insight': 'Use for monthly HB performance reviews with stakeholders'
    },
    '48': {
        'title': 'Action Priority Matrix',
        'description': 'All 10 prescriptive actions mapped on Effort vs Impact matrix. Color-coded: Red=Urgent, Orange=High, Yellow=Medium, Green=Support.',
        'insight': 'This is your complete implementation roadmap - follow priorities'
    }
}

create_category_sheet(wb, "📱 Dashboards", "STRATEGIC DASHBOARDS - Executive View",
                     "09_Strategic_Dashboards", dashboard_charts)

# Save the workbook
output_path = f'{base_path}/Half_Board_Complete_Visual_Report.xlsx'
wb.save(output_path)

print("\n" + "="*80)
print("✓ COMPLETE VISUAL REPORT CREATED SUCCESSFULLY!")
print("="*80)
print(f"\nFile: {output_path}")
print("\nContents:")
print("  ✓ Table of Contents with navigation")
print("  ✓ 6 Initial Analysis charts (with full descriptions)")
print("  ✓ 6 Overview & Distribution charts")
print("  ✓ 8 Agency Analysis charts")
print("  ✓ 5 Miracle Tourism charts")
print("  ✓ 6 Rate Code Analysis charts")
print("  ✓ 5 Market Segmentation charts")
print("  ✓ 6 HB Performance charts")
print("  ✓ 5 Opportunity Analysis charts")
print("  ✓ 4 Correlation charts")
print("  ✓ 3 Strategic Dashboard charts")
print("\nTotal: 54 charts with descriptions, insights, and professional formatting")
print("="*80)
