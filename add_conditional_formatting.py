import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, IconSetRule
from openpyxl.styles.differential import DifferentialStyle

print("="*80)
print("ADDING CONDITIONAL FORMATTING TO EXCEL WORKBOOK")
print("="*80)

# Load the workbook
wb = load_workbook('/home/gee_devops254/Downloads/Half Board/Half_Board_Comprehensive_Analysis_Revised.xlsx')

print("\nApplying conditional formatting...")

# ============================================================================
# SHEET 2: Agency Deep Dive - Add conditional formatting
# ============================================================================
print("  [1/7] Agency Deep Dive...")
ws2 = wb["Agency Deep Dive - Top 20"]

# Color scale for % HB Nights (column E)
ws2.conditional_formatting.add(
    'E2:E21',
    ColorScaleRule(
        start_type='min', start_color='F8696B',  # Red for low
        mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow for medium
        end_type='max', end_color='63BE7B'  # Green for high
    )
)

# Color scale for % HB Revenue (column I)
ws2.conditional_formatting.add(
    'I2:I21',
    ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )
)

# Data bars for Total Room Nights (column B)
ws2.conditional_formatting.add(
    'B2:B21',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="638EC6",
        showValue=True
    )
)

# Data bars for HB Room Nights (column C)
ws2.conditional_formatting.add(
    'C2:C21',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="63BE7B",
        showValue=True
    )
)

# Data bars for Total Revenue (column F)
ws2.conditional_formatting.add(
    'F2:F21',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="9C6ADE",
        showValue=True
    )
)

# ============================================================================
# SHEET 3: Miracle Tourism - Highlight key metrics
# ============================================================================
print("  [2/7] Miracle Tourism...")
ws3 = wb["Miracle Tourism Deep Dive"]

# Highlight revenue cells
for row in [6, 9]:
    ws3[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws3[f'B{row}'].font = Font(bold=True)

# ============================================================================
# SHEET 4: Universal Rate Codes - Add data bars
# ============================================================================
print("  [3/7] Universal Rate Codes...")
ws4 = wb["Universal Rate Codes"]

# Find the TOBBWI agencies section
tobbwi_start = None
for row in range(1, ws4.max_row):
    if ws4[f'A{row}'].value == 'TOP AGENCIES - TOBBWI':
        tobbwi_start = row + 2
        break

if tobbwi_start:
    # Data bars for TOBBWI Room Nights
    ws4.conditional_formatting.add(
        f'B{tobbwi_start}:B{tobbwi_start+14}',
        DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=None,
            color="638EC6",
            showValue=True
        )
    )

    # Color scale for % HB
    ws4.conditional_formatting.add(
        f'E{tobbwi_start}:E{tobbwi_start+14}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=25, mid_color='FFEB84',
            end_type='num', end_value=50, end_color='63BE7B'
        )
    )

# Find the TOBBJN agencies section
tobbjn_start = None
for row in range(tobbwi_start if tobbwi_start else 1, ws4.max_row):
    if ws4[f'A{row}'].value == 'TOP AGENCIES - TOBBJN':
        tobbjn_start = row + 2
        break

if tobbjn_start:
    # Data bars for TOBBJN Room Nights
    ws4.conditional_formatting.add(
        f'B{tobbjn_start}:B{tobbjn_start+14}',
        DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=None,
            color="9C6ADE",
            showValue=True
        )
    )

    # Color scale for % HB
    ws4.conditional_formatting.add(
        f'E{tobbjn_start}:E{tobbjn_start+14}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=25, mid_color='FFEB84',
            end_type='num', end_value=50, end_color='63BE7B'
        )
    )

# ============================================================================
# SHEET 5: Opportunity Matrix - Icon sets and color scales
# ============================================================================
print("  [4/7] Opportunity Matrix...")
ws5 = wb["Opportunity Matrix"]

# Icon set for Priority Score (column L)
ws5.conditional_formatting.add(
    'L2:L31',
    IconSetRule(
        icon_style='3TrafficLights1',
        type='num',
        values=[0, 5, 7],
        showValue=True,
        reverse=False
    )
)

# Color scale for Current HB % (column E)
ws5.conditional_formatting.add(
    'E2:E31',
    ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )
)

# Data bars for Est. Incremental F&B Revenue (column I)
ws5.conditional_formatting.add(
    'I2:I31',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="FF6B6B",
        showValue=True
    )
)

# Data bars for Potential HB Nights (column H)
ws5.conditional_formatting.add(
    'H2:H31',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="4ECDC4",
        showValue=True
    )
)

# ============================================================================
# SHEET 6: Market Segmentation - Color scales
# ============================================================================
print("  [5/7] Market Segmentation...")
ws6 = wb["Market Segmentation"]

# Find the market performance table
market_start = 5  # Assuming it starts at row 5

# Color scale for HB Penetration
ws6.conditional_formatting.add(
    f'F{market_start}:F{market_start+10}',
    ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=25, mid_color='FFEB84',
        end_type='num', end_value=50, end_color='63BE7B'
    )
)

# Data bars for Total Revenue
ws6.conditional_formatting.add(
    f'E{market_start}:E{market_start+10}',
    DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=None,
        color="9C6ADE",
        showValue=True
    )
)

# ============================================================================
# SHEET 7: ACTION PLAN - Priority highlighting
# ============================================================================
print("  [6/7] Action Plan...")
ws7 = wb["ACTION PLAN"]

# Icon set for Priority (column A)
ws7.conditional_formatting.add(
    'A2:A15',
    IconSetRule(
        icon_style='3Symbols',
        type='num',
        values=[4, 7, 11],
        showValue=True,
        reverse=True  # Reverse so 1 gets highest priority symbol
    )
)

# ============================================================================
# SHEET 8: EDA VISUALIZATIONS - Enhance key tables
# ============================================================================
print("  [7/7] EDA Visualizations...")
ws8 = wb["EDA VISUALIZATIONS"]

# Find HB comparison table (Table 1)
hb_comp_start = 4
if ws8['A4'].value == 'Category':
    # Data bars for Revenue
    ws8.conditional_formatting.add(
        f'D5:D6',
        DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=None,
            color="4472C4",
            showValue=True
        )
    )

# Find Top 15 Agencies table
top15_start = None
for row in range(1, 50):
    if ws8[f'A{row}'].value and 'TOP 15 AGENCIES' in str(ws8[f'A{row}'].value):
        top15_start = row + 2
        break

if top15_start:
    # Data bars for HB Room Nights
    ws8.conditional_formatting.add(
        f'C{top15_start}:C{top15_start+14}',
        DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=None,
            color="63BE7B",
            showValue=True
        )
    )

    # Data bars for HB Revenue
    ws8.conditional_formatting.add(
        f'E{top15_start}:E{top15_start+14}',
        DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=None,
            color="FF6B6B",
            showValue=True
        )
    )

# Find Top 15 HB Performers table
hb_perf_start = None
for row in range(top15_start if top15_start else 1, 100):
    if ws8[f'A{row}'].value and 'TOP 15 HALF BOARD' in str(ws8[f'A{row}'].value):
        hb_perf_start = row + 2
        break

if hb_perf_start:
    # Color scale for % HB Nights
    ws8.conditional_formatting.add(
        f'D{hb_perf_start}:D{hb_perf_start+14}',
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )

# Find Booking Size Distribution table
booking_start = None
for row in range(hb_perf_start if hb_perf_start else 1, ws8.max_row):
    if ws8[f'A{row}'].value and 'BOOKING SIZE DISTRIBUTION' in str(ws8[f'A{row}'].value):
        booking_start = row + 2
        break

if booking_start:
    # Color scale for HB %
    ws8.conditional_formatting.add(
        f'D{booking_start}:D{booking_start+5}',
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )

    # Icon set for HB %
    ws8.conditional_formatting.add(
        f'D{booking_start}:D{booking_start+5}',
        IconSetRule(
            icon_style='3Arrows',
            type='percent',
            values=[33, 67],
            showValue=True
        )
    )

# Save the workbook
output_file = '/home/gee_devops254/Downloads/Half Board/Half_Board_Final_Analysis.xlsx'
wb.save(output_file)

print("\n" + "="*80)
print("✓ CONDITIONAL FORMATTING APPLIED SUCCESSFULLY!")
print("="*80)
print(f"\nFile location: {output_file}")
print("\nConditional Formatting Applied:")
print("="*80)
print("✓ SHEET 2 (Agency Deep Dive):")
print("  - Color scales for HB penetration %")
print("  - Data bars for room nights and revenue")
print("")
print("✓ SHEET 3 (Miracle Tourism):")
print("  - Highlighted key revenue metrics")
print("")
print("✓ SHEET 4 (Universal Rate Codes):")
print("  - Data bars for agency performance")
print("  - Color scales for HB adoption %")
print("")
print("✓ SHEET 5 (Opportunity Matrix):")
print("  - Traffic light icons for priority scores")
print("  - Color scales for HB penetration")
print("  - Data bars for revenue potential")
print("")
print("✓ SHEET 6 (Market Segmentation):")
print("  - Color scales for HB penetration by market")
print("  - Data bars for market revenue")
print("")
print("✓ SHEET 7 (Action Plan):")
print("  - Star icons for priority levels")
print("")
print("✓ SHEET 8 (EDA Visualizations):")
print("  - Data bars for key metrics")
print("  - Color scales for performance indicators")
print("  - Arrow icons for HB penetration trends")
print("")
print("="*80)
print("LEGEND:")
print("="*80)
print("🔴 RED = Low performance / High priority action needed")
print("🟡 YELLOW = Medium performance / Moderate attention")
print("🟢 GREEN = High performance / Well optimized")
print("")
print("📊 Data Bars = Visual comparison of values")
print("🚦 Traffic Lights = Priority indicators")
print("⭐ Stars = Action priority ranking")
print("➡️ Arrows = Trend indicators")
print("="*80)
