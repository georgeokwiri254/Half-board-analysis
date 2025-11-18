import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from pathlib import Path

print("="*80)
print("CREATING COMPREHENSIVE VISUAL EXCEL REPORT")
print("="*80)

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=16)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
description_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
description_font = Font(size=10, italic=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Base path
base_path = "/home/gee_devops254/Downloads/Half Board"

# ============================================================================
# SHEET 1: TABLE OF CONTENTS
# ============================================================================
print("\n[1/11] Creating Table of Contents...")
ws_toc = wb.create_sheet("📑 Table of Contents")

ws_toc['A1'] = 'HALF BOARD ANALYSIS - VISUAL REPORT'
ws_toc['A1'].font = Font(bold=True, size=20, color="1F4E78")
ws_toc.merge_cells('A1:D1')
ws_toc.row_dimensions[1].height = 30

ws_toc['A3'] = 'COMPREHENSIVE VISUAL ANALYSIS WITH 54 CHARTS & INSIGHTS'
ws_toc['A3'].font = Font(bold=True, size=14, color="2C3E50")
ws_toc.merge_cells('A3:D3')

ws_toc['A5'] = 'Sheet'
ws_toc['B5'] = 'Category'
ws_toc['C5'] = 'Charts'
ws_toc['D5'] = 'Description'

for col in ['A', 'B', 'C', 'D']:
    ws_toc[f'{col}5'].fill = header_fill
    ws_toc[f'{col}5'].font = header_font
    ws_toc[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
    ws_toc[f'{col}5'].border = border

toc_data = [
    ['Sheet 2', 'Initial Analysis Suite', '6 charts', 'Univariate, Multivariate, CIS, Correlation, Dashboard'],
    ['Sheet 3', 'Overview & Distribution', '6 charts', 'HB penetration, revenue split, rate comparisons'],
    ['Sheet 4', 'Agency Analysis', '8 charts', 'Top performers, HB adoption, booking patterns'],
    ['Sheet 5', 'Miracle Tourism Deep Dive', '5 charts', 'Why #1 in HB, Luxembourg market insights'],
    ['Sheet 6', 'Rate Code Analysis', '6 charts', 'Universal codes, TOBBWI & TOBBJN opportunities'],
    ['Sheet 7', 'Market Segmentation', '5 charts', 'CIS, Luxembourg, Universal market performance'],
    ['Sheet 8', 'HB Performance', '6 charts', 'Pareto, heatmaps, penetration patterns'],
    ['Sheet 9', 'Opportunity Analysis', '5 charts', 'Priority matrix, AED 920k potential'],
    ['Sheet 10', 'Correlations', '4 charts', 'Statistical relationships, matrices'],
    ['Sheet 11', 'Strategic Dashboards', '3 charts', 'Executive summary, scorecard, priorities'],
]

row = 6
for data in toc_data:
    for col_idx, value in enumerate(data, 1):
        cell = ws_toc.cell(row, col_idx, value)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 1:
            cell.font = Font(bold=True)
    row += 1

# Key Statistics Summary
ws_toc[f'A{row+2}'] = 'KEY STATISTICS'
ws_toc[f'A{row+2}'].font = Font(bold=True, size=12, color="1F4E78")
ws_toc.merge_cells(f'A{row+2}:B{row+2}')

stats_data = [
    ['Total Charts:', '54'],
    ['HB Penetration:', '14% (86% opportunity)'],
    ['Incremental Revenue Potential:', 'AED 920,000'],
    ['Top HB Agency:', 'Miracle Tourism (151 nights)'],
    ['Best Market:', 'CIS (36% HB penetration)'],
    ['Universal Code Opportunity:', '5,445 room nights'],
]

row = row + 3
for stat in stats_data:
    ws_toc[f'A{row}'] = stat[0]
    ws_toc[f'B{row}'] = stat[1]
    ws_toc[f'A{row}'].font = Font(bold=True)
    ws_toc[f'B{row}'].font = Font(color="E74C3C", bold=True)
    row += 1

ws_toc.column_dimensions['A'].width = 15
ws_toc.column_dimensions['B'].width = 30
ws_toc.column_dimensions['C'].width = 15
ws_toc.column_dimensions['D'].width = 50

# ============================================================================
# SHEET 2: INITIAL ANALYSIS SUITE (6 charts)
# ============================================================================
print("[2/11] Creating Initial Analysis Suite...")
ws_initial = wb.create_sheet("📊 Initial Analysis Suite")

ws_initial['A1'] = 'INITIAL STATISTICAL ANALYSIS - 6 KEY VISUALIZATIONS'
ws_initial['A1'].font = title_font
ws_initial['A1'].fill = title_fill
ws_initial['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_initial.merge_cells('A1:H1')
ws_initial.row_dimensions[1].height = 30

initial_charts = [
    {
        'path': f'{base_path}/1_univariate_analysis.png',
        'title': 'Chart 1: Univariate Analysis',
        'description': 'Distribution analysis of key metrics: Room Nights, Revenue, and Average Rates. Shows data spread, outliers (box plots), and central tendencies. Key insight: High variance in bookings with extreme outliers.',
        'row': 3
    },
    {
        'path': f'{base_path}/2_halfboard_analysis.png',
        'title': 'Chart 2: Half Board Analysis',
        'description': 'HB vs Non-HB comparison across agencies. Top 10 agencies by HB performance showing Miracle Tourism leading with 151 nights. Visual gap demonstrates conversion opportunity.',
        'row': 25
    },
    {
        'path': f'{base_path}/3_multivariate_analysis.png',
        'title': 'Chart 3: Multivariate Analysis',
        'description': 'Agency & Rate Code relationships. Scatter plots and performance matrices showing correlation between volume and HB adoption. Top agency-rate combinations identified.',
        'row': 47
    },
    {
        'path': f'{base_path}/4_cis_market_analysis.png',
        'title': 'Chart 4: CIS Market Analysis',
        'description': 'CIS markets show 36% HB penetration (best in dataset) but only 9 bookings total. TOKHACIS leads with AED 17,888 revenue. Massive scale opportunity identified.',
        'row': 69
    },
    {
        'path': f'{base_path}/5_correlation_heatmap.png',
        'title': 'Chart 5: Correlation Heatmap',
        'description': 'Pearson correlation between Room Nights, Revenue, and Avg Rate. Perfect 0.999 correlation between nights and revenue. Rate shows weak negative correlation with volume.',
        'row': 91
    },
    {
        'path': f'{base_path}/6_comprehensive_dashboard.png',
        'title': 'Chart 6: Comprehensive Dashboard',
        'description': 'Multi-panel overview: Key KPIs, revenue distribution, top agencies, booking patterns, and rate comparisons. Single-page executive summary of entire analysis.',
        'row': 113
    }
]

for chart_info in initial_charts:
    row = chart_info['row']

    # Title
    ws_initial[f'A{row}'] = chart_info['title']
    ws_initial[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")
    ws_initial.merge_cells(f'A{row}:H{row}')
    ws_initial.row_dimensions[row].height = 25

    # Description
    ws_initial[f'A{row+1}'] = chart_info['description']
    ws_initial[f'A{row+1}'].font = description_font
    ws_initial[f'A{row+1}'].fill = description_fill
    ws_initial[f'A{row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_initial.merge_cells(f'A{row+1}:H{row+2}')
    ws_initial.row_dimensions[row+1].height = 40

    # Image
    if os.path.exists(chart_info['path']):
        img = Image(chart_info['path'])
        # Resize to fit (width ~1000 pixels = ~14 columns)
        img.width = 1000
        img.height = int(1000 * img.height / img.width) if img.width > 0 else 400
        ws_initial.add_image(img, f'A{row+3}')
        # Set row height to accommodate image
        for i in range(20):
            ws_initial.row_dimensions[row+3+i].height = 20

ws_initial.column_dimensions['A'].width = 15

# ============================================================================
# HELPER FUNCTION TO ADD CHART CATEGORY SHEETS
# ============================================================================
def create_category_sheet(wb, sheet_name, title, chart_folder, chart_descriptions):
    """Create a sheet with charts from a specific category"""
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    current_row = 3

    for chart_num, desc in chart_descriptions.items():
        # Find chart file
        chart_files = list(Path(f'{base_path}/charts/{chart_folder}').glob(f'{chart_num}_*.png'))

        if not chart_files:
            continue

        chart_path = str(chart_files[0])

        # Title
        ws[f'A{current_row}'] = f'Chart {chart_num}: {desc["title"]}'
        ws[f'A{current_row}'].font = Font(bold=True, size=13, color="2C3E50")
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws.row_dimensions[current_row].height = 25

        # Description
        ws[f'A{current_row+1}'] = desc['description']
        ws[f'A{current_row+1}'].font = description_font
        ws[f'A{current_row+1}'].fill = description_fill
        ws[f'A{current_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row+1}:H{current_row+2}')
        ws.row_dimensions[current_row+1].height = 35

        # Key Insight box
        ws[f'A{current_row+3}'] = '🔑 KEY INSIGHT:'
        ws[f'B{current_row+3}'] = desc['insight']
        ws[f'A{current_row+3}'].font = Font(bold=True, color="E74C3C")
        ws[f'B{current_row+3}'].font = Font(bold=True, color="27AE60")
        ws[f'B{current_row+3}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        ws.merge_cells(f'B{current_row+3}:H{current_row+3}')
        ws.row_dimensions[current_row+3].height = 20

        # Image
        if os.path.exists(chart_path):
            img = Image(chart_path)
            img.width = 950
            img.height = int(950 * img.height / img.width) if img.width > 0 else 380
            ws.add_image(img, f'A{current_row+4}')

            # Set row heights for image
            rows_needed = int(img.height / 20) + 1
            for i in range(rows_needed):
                ws.row_dimensions[current_row+4+i].height = 20

            current_row += rows_needed + 6
        else:
            current_row += 25

    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14

# ============================================================================
# SHEET 3: OVERVIEW & DISTRIBUTION
# ============================================================================
print("[3/11] Creating Overview & Distribution...")
overview_charts = {
    '01': {
        'title': 'Overall HB Penetration',
        'description': 'Pie chart showing only 14% of bookings include Half Board, leaving 86% as opportunity gap. This is the most critical metric - immediate action required.',
        'insight': 'CRITICAL: 86% of bookings have NO F&B packages!'
    },
    '02': {
        'title': 'Revenue Distribution',
        'description': 'HB generates only 3.1% of total revenue despite being 14% of bookings. This indicates HB bookings are undermonetized.',
        'insight': 'HB is underperforming in revenue generation - pricing issue'
    },
    '03': {
        'title': 'Room Nights Distribution',
        'description': 'Only 2.7% of room nights include F&B packages (648 out of 23,747 nights). Massive untapped potential.',
        'insight': '97.3% of room nights lack F&B attachment'
    },
    '04': {
        'title': 'Booking Count Distribution',
        'description': 'Visual comparison: 50 HB bookings vs 307 non-HB bookings. Bar chart emphasizes the conversion opportunity.',
        'insight': 'Convert non-HB to HB through incentives and bundling'
    },
    '05': {
        'title': 'Average Rate Comparison',
        'description': 'HB average rate (AED 419.30) only AED 26.21 higher than non-HB (AED 393.09). Premium is insufficient.',
        'insight': 'Increase HB surcharge to AED 120-150 per night'
    },
    '06': {
        'title': 'Revenue vs Room Nights Scatter',
        'description': 'Scatter plot of all bookings showing perfect linear relationship (0.999 correlation). Green = HB, Red = Non-HB.',
        'insight': 'Strong correlation confirms: More room nights = More revenue'
    }
}

create_category_sheet(wb, "📈 Overview", "OVERVIEW & DISTRIBUTION ANALYSIS",
                     "01_Overview_Distribution", overview_charts)

# ============================================================================
# SHEET 4: AGENCY ANALYSIS
# ============================================================================
print("[4/11] Creating Agency Analysis...")
agency_charts = {
    '07': {
        'title': 'Top 20 Agencies by Revenue',
        'description': 'Revenue leaders ranked. "Totals" dominates with AED 3.8M. TBO, Darina, Webbeds follow. Color gradient shows relative performance.',
        'insight': 'Top 5 agencies generate 60%+ of revenue - prioritize these'
    },
    '08': {
        'title': 'Top 20 Agencies by Room Nights',
        'description': 'Volume leaders: "Totals" (11,874 nights), TBO (758), Darina (932). High volume = high conversion potential.',
        'insight': 'Target these high-volume agencies for HB conversion'
    },
    '09': {
        'title': 'Top 15 HB Agencies by Nights',
        'description': 'Miracle Tourism leads HB performance with 151 nights - 2x more than #2 (Desert Gate: 71 nights). Clear winner.',
        'insight': 'Miracle Tourism is the HB champion - study their model!'
    },
    '10': {
        'title': 'Top 15 HB Agencies by Revenue',
        'description': 'Miracle generates AED 55,130 in HB revenue alone. Al Khalidiah (AED 20,115) and Voyage (AED 19,188) follow.',
        'insight': 'Create "HB Champions" program modeled after Miracle'
    },
    '11': {
        'title': 'HB Penetration by Top 20',
        'description': 'Color-coded: Green >25%, Orange 10-25%, Red <10%. Most top agencies in RED zone. Target line at 35%.',
        'insight': 'Most top agencies have <10% HB - these are HIGH PRIORITY targets'
    },
    '12': {
        'title': 'HB vs Non-HB Stacked',
        'description': 'Stacked bars showing HB (green) vs Non-HB (red) split for top 15 agencies. Visual gap emphasizes opportunity.',
        'insight': 'Use this chart in agency presentations to show opportunity'
    },
    '13': {
        'title': 'Average Rate Comparison',
        'description': 'Grouped bars comparing HB vs Non-HB rates by agency. Some agencies have LOWER HB rates (pricing error!).',
        'insight': 'Standardize HB premium at +AED 120 across all agencies'
    },
    '14': {
        'title': 'Booking Size Box Plot',
        'description': 'Box plots showing booking size variance by agency. Outliers indicate large group bookings with high HB potential.',
        'insight': 'Target agencies with large bookings for mandatory HB'
    }
}

create_category_sheet(wb, "🏢 Agencies", "AGENCY ANALYSIS",
                     "02_Agency_Analysis", agency_charts)

# ============================================================================
# SHEET 5: MIRACLE TOURISM DEEP DIVE
# ============================================================================
print("[5/11] Creating Miracle Tourism Deep Dive...")
miracle_charts = {
    '15': {
        'title': 'Miracle vs Top 5 Comparison',
        'description': 'Normalized multi-metric comparison (0-100 scale). Miracle excels in HB penetration despite moderate volume.',
        'insight': "Miracle's model is the blueprint - apply to other agencies"
    },
    '16': {
        'title': 'Miracle Booking Size Distribution',
        'description': 'Histogram showing most Miracle bookings are 30+ nights (large groups/packages). This drives HB success.',
        'insight': 'Large bookings = Higher HB adoption. Target long-stay agencies.'
    },
    '17': {
        'title': 'Miracle HB vs Non-HB Split',
        'description': 'Pie chart showing Miracle has achieved high HB penetration in their portfolio (compared to 14% overall).',
        'insight': 'Showcase Miracle as success story in all agency meetings'
    },
    '18': {
        'title': 'Miracle Rate Code Performance',
        'description': 'TOMILUX (Luxembourg market rate code) drives majority of Miracle revenue. Specialized market strategy.',
        'insight': 'Luxembourg market guests prefer packages - expand to similar markets'
    },
    '19': {
        'title': 'Miracle Revenue Contribution',
        'description': 'Single agency contributing significant share of total HB revenue. Pie chart shows concentration risk.',
        'insight': 'Reduce dependency by scaling HB across other agencies'
    }
}

create_category_sheet(wb, "⭐ Miracle", "MIRACLE TOURISM DEEP DIVE (Luxembourg Market)",
                     "03_Miracle_Deep_Dive", miracle_charts)

# Continue creating remaining sheets...
print("[6/11] Creating Rate Code Analysis...")
# (continuing in next part due to length...)

print("\n" + "="*80)
print("Creating remaining category sheets...")
print("="*80)
